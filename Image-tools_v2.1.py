import time
nowtimer = time.time()
print(nowtimer)
import psutil
import functools
import tempfile
import logging
import os
import re
import shutil
import sys
import threading
import tkinter as tk
from collections import defaultdict
from pathlib import Path
from tkinter import scrolledtext
from tkinter import messagebox, filedialog,simpledialog # 删除ttk
from typing import Dict, List #, Any, Optional, Tuple
import configparser
import cv2
import types
import fitz  # PyMuPDF
import numpy as np
import requests
import datetime
from PIL import Image, ImageTk
from openpyxl import Workbook
from packaging.version import parse as parse_version
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
import ttkbootstrap as ttk
from ttkbootstrap.constants import *
import paddle
print(time.time()-nowtimer,"**PaddleOCR** start import")
from paddleocr import PaddleOCR
print(time.time()-nowtimer,"**PaddleOCR** stop import")
# Set environment variables to ensure dynamic library loading
os.environ['PATH'] += os.pathsep + os.path.dirname(sys.executable)

def resource_path(relative_path):
    """Get resource path after packaging"""
    try:
        base_path = sys._MEIPASS  # PyInstaller temporary directory
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.normpath(os.path.join(base_path, relative_path))

# Configure logging
logging.basicConfig(
    level=logging.WARNING, # DEBUG
    format='%(asctime)s - %(levelname)s - %(message)s',
    filename=resource_path('Image-tools.log')
)

# Add PaddlePaddle DLL path
if getattr(sys, 'frozen', False):
    bundle_dir = sys._MEIPASS
    os.environ['PATH'] += os.pathsep + os.path.join(bundle_dir, 'paddle', 'libs')


# -------------------- Language Support --------------------
languages = {
    "zh_CN": {
        # 主界面
        "title": "图片处理工具 v2.1",
        "tab_recognize": "识别文字/发票",
        "tab_convert": "转换与合并",
        "tab_optimize": "优化图片",
        "tab_settings": "设置",
        "tab_about": "关于",
        "use_gpu": "启用GPU加速",
        "保留识别内容": "保留历史识别内容",
        # 新增翻译
        "compute_precision": "计算精度",
        "fp32_mode": "单精度(FP32)",
        "fp16_mode": "半精度(FP16)",

        "current_mode": "当前模式",  # 新增
        "pan_mode": "平移模式",
        "crop_mode": "裁剪模式",
        "zoom_tips": "缩放提示：\n1. 使用鼠标滚轮缩放图片\n2. 按住左键拖动查看细节\n3. 双击图片重置缩放",
        # 识别标签页
        "file_path": "文件路径:",
        "browse": "浏览",
        "settings": "设置选项",
        "excel_name": "Excel名称:",
        "recognition_optimize": "识别优化:",
        "ocr_module": "OCR模块:",
        "start_recognition": "开始识别",
        "export_to_a1": "导出到A1",
        "row": "行:",
        "column": "列:",
        "export_to_position": "指定位置导出",
        "recognition_result": "识别结果",
        "copy_content": "复制内容",
        "clear_content": "清空内容",
        "invoice_key_info": "【发票】关键信息",
        "field": "字段",
        "value": "值",
        "file_preview": "文件预览",

        # 优化标签页
        "select_image": "选择图片:",
        "optimization_options": "优化选项",
        "denoise_strength": "去噪强度(0~20)",
        "sharpen_strength": "锐化强度(0~5)",
        "brightness": "亮度(-100~100)",
        "contrast": "对比度(-100~100)",
        "edge_enhancement": "边缘增强(0~5)",
        "save_result": "保存结果",
        "reset": "重置",
        "original_preview": "原图预览",
        "processed_preview": "处理后预览",

        # 转换标签页
        "select_conversion": "选择转换类型:",
        "pdf_to_image": "PDF转PNG/JPG/TIFF",
        "image_to_pdf": "图片转PDF",
        "txt_to_pdf": "TXT转PDF",
        "merge_pdf": "PDF合并",
        "start_conversion": "开始转换",
        "merge_multiple_images": "合并多张图片",
        "browse_txt": "浏览TXT",
        "add_file": "添加文件",
        "delete_selected": "删除选中",
        "start_merge": "开始合并",
        "conversion_result": "转换结果",

        # 关于标签页
        "developer": "开发者:TangKris(小汤)",
        "dev_date": "开发日期:2025.03.02 - 发布日期:2025.3.15",
        "current_version": "当前程序版本  v{version}",
        "check_update": "检查更新",
        "update_log": "更新日志",
        "copyright": "©TangKris 版权所有",

        # 语言选择
        "language": "语言选择:",
        "lang_zh": "简体中文",
        "lang_en": "English",
        "apply_changes": "应用修改",
        "loading": "加载中...",
        "theme_switch": "界面风格:",
        "style_minty": "薄荷绿",
        "style_darkly": "暗黑",
        "style_solar": "太阳能",
        "style_superhero": "超级英雄",
        # 消息
        "success": "成功",
        "error": "错误",
        "warning": "警告",
        "info": "提示",
        "recognition_complete": "识别完成！",
        "recognition_failed": "识别失败:",
        "excel_saved": "已生成并保存Excel！",
        "excel_save_failed": "Excel保存失败！",
        "empty_file_path": "文件路径不能为空",
        "empty_excel_name": "未填写表格名！",
        "copied_to_clipboard": "内容已复制到剪贴板！",
        "nothing_to_copy": "没有内容可复制！",
        "invalid_row_col": "无效的行列输入: {error}",
        "load_image_failed": "无法加载图片: {error}",
        "optimization_failed": "优化处理失败: {error}",
        "image_saved": "图片保存成功！",
        "save_failed": "保存失败: {error}",
        "select_image_first": "请先选择图片",
        "input_range_error": "输入值必须在 {min} 到 {max} 之间",
        "input_value_error": "输入值不能包含字母、特殊符号",
        "error_code": "错误代码:{error}",
        "no_image_selected": "未选择图片",
        "conversion_complete": "{type}转换完成！",
        "conversion_failed": "{type}转换失败：{error}",
        "pdf_converted": "PDF文件已成功转换为{format}, 保存在:\n{path}",
        "images_converted": "已将 {count} 张图片成功转换为 PDF 文件，保存为：\n{path}",
        "txt_converted": "已将TXT文件成功转换为PDF，保存为：\n{path}",
        "pdf_merged": "PDF文件合并完成，保存为:\n{path}",
        "merge_failed": "合并失败：{error}",
        "no_update": "无更新",
        "no_valid_version": "未找到有效版本",
        "already_latest": "当前已是最新版本 (v{version})",
        "network_error": "连接仓库失败: {error}",
        "update_error": "检查更新失败: {error}",
        "update_available": "发现新版本 v{version}",
        "update_available_tips": "发现新版本 v{version} ({count}个分卷)\n是否立即下载？",
        "no_valid_parts": "未找到有效分卷文件",
        "download_part": "下载分卷 {current}/{total}",
        "downloading": "正在下载: {name}",
        "download_failed": "{name} 下载失败: {error}",
        "download_complete": "文件已保存到: {path}\n请手动解压替换旧版本",
        "language_changed": "语言已更改为中文",

        # 优化选项
        "optimize_none": "无",
        "optimize_grayscale": "灰度增强",
        "optimize_binary": "二值化",
        "optimize_denoise": "去噪点",
        "optimize_grayscale_binary": "灰度增强和二值化",
        "optimize_binary_denoise": "二值化和去噪点",
        "optimize_all": "全部启用",

        # OCR模块
        "ocr_paddle": "PaddleOCR",

        # 转换选项
        "conversion_all": "全部",
        "conversion_pdf_to_image": "PDF转PNG/JPG/TIFF",
        "conversion_image_to_pdf": "图片转PDF",
        "conversion_txt_to_pdf": "TXT转PDF",
        "conversion_merge_pdf": "Image合并PDF Multi",
        "conversion_info_only": "仅信息",

        # 图片和文件类型
        "image": "图片",
        "all_files": "所有文件",
        "image_pdf": "图片和PDF文件",

        "tab_edit": "编辑图片",
        "edit_select_image": "选择图片：",
        "resize_settings": "尺寸调整",
        "width": "宽度：",
        "height": "高度：",
        "keep_ratio": "保持比例",
        "save_edit": "保存修改",
        "preview_title": "编辑预览",
        "invalid_dimension": "请输入有效的数字尺寸",
        "current_size": "原图尺寸：{}x{}",
        "saving_status": "保存中...",
        "invoice_valid": "✅ 发票有效",
        "aspect_ratio": "宽高比例:",
        "preset_ratios": {
            "original": "原图比例",
            "3:2": "3:2 (照片)",
            "4:3": "4:3 (屏幕)",
            "16:9": "16:9 (视频)",
            "1:1": "1:1 (正方形)",
            "custom": "自定义比例"
        },
        "custom_ratio_prompt": "请输入宽高比（格式如 3:2）",
        "invalid_ratio_format": "无效的比例格式，请使用'数字:数字'格式",
        "crop_settings": "图片裁剪",
        "start_crop": "开始裁剪",
        "crop_coords": "裁剪坐标:",
        "crop_error": "请先选择裁剪区域",
        "invalid_crop": "无效的裁剪区域",
        "crop_tips": "裁剪提示：\n1. 在图片上点击并拖动以选择裁剪区域\n2. 松开鼠标后会自动调整裁剪区域到有效范围\n3. 裁剪区域会实时预览\n4. 点击'保存修改'按钮保存裁剪后的图片",
        "invoice_duplicate_check": "发票重复识别",
        "save_settings": "保存设置",
        "output_format": "输出格式",
        "no_invoice_info": "没有可保存的发票信息",
        "invoice_saved": "发票信息已保存",
        "clear_selected": "清除当前发票记录",
        "clear_all": "清除全部发票记录",
        "validation_result": "验证结果",
        "verification_time": "验证时间",
        "history_cleared": "已清除{count}条历史记录",
        "selected_cleared": "已清除选定记录",
        "no_records_selected": "未识别到发票数据",
        "warning_prompt": "检测到重复发票时弹出警告提示",
        "enable_invoice_recognition": "启用发票识别",
        "identify_recognition": "识别设置",
        "CPU_throttling_settings": "CPU节流设置",
        "Identify_acceleration_options": "识别加速选项",
        "CPU_max%use_frame":"CPU使用率阈值(%):",
        "CPU_tips_frame":"(超过此值将自动降速)",
        "computer_into":"计算精度:",
        "computer_max_into": "计算精度支持: 检测中...",
        "setting_tips": "部分功能需重新启动程序才能完全生效",
        "setting_tips(d)": "请勿随意修改配置文件，否则可能引起程序报错！",
    },
    "en_US": {
        # 主界面
        "title": "Image Processing Tool v2.1",
        "tab_recognize": "Recognize Text/Invoice",
        "tab_convert": "Convert & Merge",
        "tab_optimize": "Optimize Image",
        "tab_about": "About",
        "use_gpu": "Enable GPU Acceleration",
        "recognition_complete": "Recognition complete! Processed {total} pages",
        "accel_settings": "Hardware Acceleration",
        "intel_mkldnn": "Enable Intel MKLDNN",
        "nvidia_tensorrt": "Enable NVIDIA TensorRT",
        "compute_precision": "Compute Precision",
        "保留识别内容": "Retain recognition history",
        "fp32_mode": "FP32 Precision",
        "fp16_mode": "FP16 Precision",
        # 识别标签页
        "file_path": "File Path:",
        "browse": "Browse",
        "settings": "Settings",
        "excel_name": "Excel Name:",
        "recognition_optimize": "Recognition Optimization:",
        "ocr_module": "OCR Module:",
        "start_recognition": "Start Recognition",
        "export_to_a1": "Export to A1",
        "row": "Row:",
        "column": "Column:",
        "export_to_position": "Export to Position",
        "recognition_result": "Recognition Result",
        "copy_content": "Copy Content",
        "clear_content": "Clear Content",
        "invoice_key_info": "Invoice Key Information",
        "field": "Field",
        "value": "Value",
        "file_preview": "File Preview",

        "current_mode": "Current Mode",  # 新增
        "pan_mode": "Pan Mode",
        "crop_mode": "Crop Mode",
        "zoom_tips": "Zoom Tips:\n1. Use mouse wheel to zoom\n2. Drag with left button to pan\n3. Double-click to reset zoom",
        # 优化标签页
        "select_image": "Select Image:",
        "optimization_options": "Optimization Options",
        "denoise_strength": "Denoise Strength (0~20)",
        "sharpen_strength": "Sharpen Strength (0~5)",
        "brightness": "Brightness (-100~100)",
        "contrast": "Contrast (-100~100)",
        "edge_enhancement": "Edge Enhancement (0~5)",
        "save_result": "Save Result",
        "reset": "Reset",
        "original_preview": "Original Preview",
        "processed_preview": "Processed Preview",

        # 转换标签页
        "select_conversion": "Select Conversion Type:",
        "pdf_to_image": "PDF to PNG/JPG/TIFF",
        "image_to_pdf": "Image to PDF",
        "txt_to_pdf": "TXT to PDF",
        "merge_pdf": "Merge PDF",
        "start_conversion": "Start Conversion",
        "merge_multiple_images": "Merge Multiple Images",
        "browse_txt": "Browse TXT",
        "add_file": "Add File",
        "delete_selected": "Delete Selected",
        "start_merge": "Start Merge",
        "conversion_result": "Conversion Result",

        # 关于标签页
        "developer": "Developer: TangKris",
        "dev_date": "Development: 2025.03.02 - Release: 2025.3.15",
        "current_version": "Current Version: v{version}",
        "check_update": "Check for Updates",
        "update_log": "Update Log",
        "copyright": "© TangKris All Rights Reserved",

        # 语言选择
        "lang_zh": "Chinese",
        "lang_en": "English",
        "apply_changes": "Apply Changes",
        "loading": "Loading...",
        "theme_switch": "Interface Theme:",
        "style_minty": "Minty",
        "style_darkly": "Darkly",
        "style_solar": "Solar",
        "style_superhero": "Superhero",
        # 消息
        "success": "Success",
        "error": "Error",
        "warning": "Warning",
        "info": "Info",
        "recognition_failed": "Recognition Failed:",
        "excel_saved": "Excel generated and saved!",
        "excel_save_failed": "Excel save failed!",
        "empty_file_path": "File path cannot be empty",
        "empty_excel_name": "Excel name is not specified!",
        "copied_to_clipboard": "Content copied to clipboard!",
        "nothing_to_copy": "Nothing to copy!",
        "invalid_row_col": "Invalid row/column input: {error}",
        "load_image_failed": "Failed to load image: {error}",
        "optimization_failed": "Optimization failed: {error}",
        "image_saved": "Image saved successfully!",
        "save_failed": "Save failed: {error}",
        "select_image_first": "Please select an image first",
        "input_range_error": "Input value must be between {min} and {max}",
        "input_value_error": "Input value cannot contain letters or special characters",
        "error_code": "Error code: {error}",
        "no_image_selected": "No image selected",
        "conversion_complete": "{type} conversion complete!",
        "conversion_failed": "{type} conversion failed: {error}",
        "pdf_converted": "PDF file successfully converted to {format}, saved at:\n{path}",
        "images_converted": "Successfully converted {count} images to PDF, saved as:\n{path}",
        "txt_converted": "Successfully converted TXT file to PDF, saved as:\n{path}",
        "pdf_merged": "PDF files merged successfully, saved at:\n{path}",
        "merge_failed": "Merge failed: {error}",
        "no_update": "No Update",
        "no_valid_version": "No valid version found",
        "already_latest": "Already on the latest version (v{version})",
        "network_error": "Connection to repository failed: {error}",
        "update_error": "Update check failed: {error}",
        "update_available": "New version v{version}",
        "update_available_tips": "New version v{version} found ({count} parts)\nDownload now?",
        "no_valid_parts": "No valid parts found",
        "download_part": "Downloading Part {current}/{total}",
        "downloading": "Downloading: {name}",
        "download_failed": "{name} download failed: {error}",
        "download_complete": "File saved to: {path}\nPlease manually extract and replace the old version",
        "language_changed": "Language changed to English",

        # 优化选项
        "optimize_none": "None",
        "optimize_grayscale": "Grayscale Enhancement",
        "optimize_binary": "Binarization",
        "optimize_denoise": "Denoising",
        "optimize_grayscale_binary": "Grayscale & Binarization",
        "optimize_binary_denoise": "Binarization & Denoising",
        "optimize_all": "Enable All",

        # OCR模块
        "ocr_paddle": "PaddleOCR (Recommended, High Accuracy, Slower)",
        "ocr_tesseract": "Tesseract (Not Recommended, Lower Accuracy, Faster)",

        # 转换选项
        "conversion_all": "All",
        "conversion_pdf_to_image": "PDF to PNG/JPG/TIFF",
        "conversion_image_to_pdf": "Image to PDF",
        "conversion_txt_to_pdf": "TXT to PDF",
        "conversion_merge_pdf": "Merge Images to PDF",
        "conversion_info_only": "Info Only",

        # 图片和文件类型
        "image": "Image",
        "all_files": "All Files",
        "image_pdf": "Images and PDF Files",

        "edit_select_image": "Select Image:",
        "resize_settings": "Resize Settings",
        "width": "Width:",
        "height": "Height:",
        "keep_ratio": "Keep Ratio",
        "save_edit": "Save Changes",
        "preview_title": "Edit Preview",
        "invalid_dimension": "Please enter valid numeric dimensions",
        "current_size": "Original Size: {}x{}",
        "saving_status": "Saving...",
        "invoice_valid": "✅ Invoice valid",
        "aspect_ratio": "Aspect Ratio:",
        "preset_ratios": {
            "original": "Original Ratio",
            "3:2": "3:2 (Photo)",
            "4:3": "4:3 (Screen)",
            "16:9": "16:9 (Video)",
            "1:1": "1:1 (Square)",
            "custom": "Custom Ratio"
        },
        "invoice_duplicate_check": "Invoice Duplicate Check",
        "save_settings": "Save Settings",
        # 修改原语言相关键值
        "custom_ratio_prompt": "Please enter ratio (e.g. 3:2)",
        "invalid_ratio_format": "Invalid ratio format, use 'number:number'",
        "crop_settings": "Image Cropping",
        "start_crop": "Start Crop",
        "crop_coords": "Coordinates:",
        "crop_error": "Please select crop area first",
        "invalid_crop": "Invalid crop area",
        "crop_tips": "Crop Tips:\n1. Click and drag on the image to select the crop area\n2. The crop area will automatically adjust to a valid range when you release the mouse\n3. The crop area will be previewed in real-time\n4. Click the 'Save Changes' button to save the cropped image",
        "output_format": "Output Format",
        "no_invoice_info": "No invoice information to save",
        "invoice_saved": "Invoice information saved",
        "clear_selected": "Clear current invoice record",
        "clear_all": "Clear All Invoice record",
        "verification_time": "Verification Time",
        "history_cleared": "Cleared {count} history records",
        "selected_cleared": "Selected records cleared",
        "validation_result": "Validation Result",
        "no_records_selected": "Invoice data not recognized",
        "warning_prompt": "Show warning popup when duplicate invoice detected",
        "enable_invoice_recognition": "Enable Invoice Recognition",
        "identify_recognition": "Identify settings",
        "CPU_throttling_settings": "CPU throttling settings",
        "Identify_acceleration_options": "Identify_acceleration_options",
        "CPU_max%use_frame": "CPU usage threshold (%):",
        "CPU_tips_frame": "(Exceeding this value will automatically slow down)",
        "computer_into": "Calculation accuracy:",
        "computer_max_into": "Calculation accuracy support: During detection",
        "setting_tips": "Some functions require a restart of the program to fully take effect",
        "setting_tips(d)": "Please do not modify the configuration file arbitrarily, otherwise it may cause program errors!",
    }
}
# -------------------- Update Log --------------------
ch_log = [
    "Image-tools Log",
    "[内测]版本v1.0:\n"
    "- 实现基础paddleocr识别功能\n"
    "- 新增创建Excel写入功能",
    "[内测]版本v1.1:\n"
    "- 实现PaddleOCR识别\n"
    "- 新增创建Excel并写入指定单元格功能\n"
    "- 无需用户手动安装OCR库",
    "[内测]版本v1.2:\n"
    "- 新增一键复制功能\n"
    "- 新增一键清空文本功能\n"
    "- 新增'转换与合并'功能，支持pdf转多种图片,多张图片转pdf，txt转pdf，pdf合并等功能\n"
    "- 程序图标换新",
    "[内测]版本v1.3:\n"
    "- 新增发票自动识别功能\n"
    "- 快速获取发票键值（发票号码等），实现初步判断真伪\n"
    "- 识别图片支持图片实时预览",
    "[内测]版本v1.4:\n"
    "- 程序窗口图标换新\n"
    "- 完善'关于'的显示内容\n"
    "- 新增'优化图片'选项卡，能够调节锐化，对比度，亮度等参数实现图片优化，支持图片预览",
    "[Debug]版本v1.5:\n"
    "- 优化图片功能增强，实时预览修改结果和数值\n"
    "- 新增'检查更新'功能，快捷更新新版本\n"
    "- 更新日志、选项卡优化",
    "[Debug]版本v1.6:\n"
    "- 修复已知bug\n"
    "- 新增语言切换功能\n"
    "- 大幅优化性能\n",
    "[Debug]版本v1.7:\n"
    "- 修复已知bug\n"
    "- 完善语言切换功能\n"
    "- 新增图片编辑功能，实现图片缩放、裁剪功能\n"
    "- 添加确认修改按钮",
    "[Debug]版本v1.8:\n"
    "- 修复已知bug，删除部分组件\n"
    "- 新增发票记录功能，支持清除记录\n"
    "- 支持永久保存语言",
    "[Debug]版本v1.9:\n"
    "- 大幅度提升程序启动速度\n"
    "- 完善发票记录功能\n"
    "- 修复语言翻译缺陷",
    "[Debug]版本v2.0:\n"
    "- 大幅度美化UI，提升用户体验\n"
    "- 提升加载速度\n"
    "- 新增风格切换功能\n"
    "- 新增编辑图片可拖动、缩放功能\n"
    "- 支持兼容Win7/10/11 MacOS Liunx系统\n",
    "[Debug]版本v2.1:\n"
    "- 新增多方面识别优化设置\n"
    "- 提升初始化速度\n"
    "- 增强发票识别用户体验",
]
en_log = [
    "Image-tools Log",
    "[Beta] Version 1.0:\n"
    "- Implemented basic PaddleOCR recognition function\n"
    "- Added function to create and write to Excel",
    "[Beta] Version 1.1:\n"
    "- Implemented PaddleOCR recognition\n"
    "- Added function to create Excel and write to specified cells\n"
    "- No need for users to manually install OCR libraries",
    "[Beta] Version 1.2:\n"
    "- Added one-click copy function\n"
    "- Added one-click clear text function\n"
    "- Added 'Convert & Merge' function, supporting PDF to multiple image formats, multiple images to PDF, TXT to PDF, and PDF merging\n"
    "- Updated program icon",
    "[Beta] Version 1.3:\n"
    "- Added automatic invoice recognition function\n"
    "- Quickly obtain invoice key values (invoice code, number, check code, etc.) to make preliminary authenticity judgments\n"
    "- Added real-time preview for recognized images",
    "[Beta] Version 1.4:\n"
    "- Updated program window icon\n"
    "- Improved content displayed in the 'About' section\n"
    "- Added 'Optimize Image' tab, allowing adjustment of parameters such as sharpening, contrast, and brightness for image optimization with preview",
    "[Debug] Version 1.5:\n"
    "- Enhanced image optimization function with real-time preview of modifications and values\n"
    "- Added 'Check for Updates' function for quick updates to new versions\n"
    "- Updated changelog and optimized tabs",
    "[Debug] Version 1.6:\n"
    "- Fixed known bugs\n"
    "- Added English-Chinese switching function\n"
    "- Significantly improved performance\n"
    "- Added confirmation button for modifications",
    "[Debug] version 1.7:\n"
    "- fix known bugs \n"
    "- improve language switching \n"
    "- add image editing function to realize image zooming and cropping \n"
    "- new picture editing function \n",
    "[Debug] version v1.8:\n"
    "- fix known bugs and delete some components \n"
    "- new invoice record function, support clearing records \n"
    "- support permanent language saving and repair language translation defects",
    "[Debug] version v1.9:\n"
    "- greatly increase program startup speed \n"
    "- improve invoice recording \n"
    "- fix language translation defects",
    "[Debug] version v2.0:\n"
    "- greatly beautify UI and improve user experience \n"
    "- increase loading speed \n"
    "- new style switching function\n"
    "- Supports compatibility with Win7/10/11 MacOS Liunx system\n"
    "- Enhance invoice recognition user experience",
    "[Debug] version v2.1: \n"
    "- Identification and optimization settings for newly added aspects\n"
    "- Improve initialization speed\n"
    "- Enhance the user experience of invoice recognition"
]
def extract_version(ch_log_entry):
    match = re.search(r"版本v(\d+\.\d+)", ch_log_entry)
    if match:
        return match.group(1)
    return "未知版本"
def get_key_from_value(dictionary, value):
    for key, val in dictionary.items():
        if val == value:
            return key
    return None  # 如果没有找到对应的键，返回 None
latest_version = extract_version(ch_log[-1])
print(latest_version)
class InvoiceValidator:
    def __init__(self):
        self.patterns = {
            "发票类型": r"增值税\s*([\u4e00-\u9fa5]+)发票",
            "发票号码": r"发票号码[:：]?\s*([A-Z0-9]{8,20})",
            "开票日期": r"开票日期[:：]?\s*(\d{4}年\d{1,2}月\d{1,2}日)"
        }

    def check_duplicate(self, number, history,warning_prompt_var):
        if number in history:
            if warning_prompt_var:  # 新增警告检查
                messagebox.showwarning(
                    "重复发票警告",
                    f"检测到重复发票号码：{number}\n首次验证时间：{history[number]}"
                )
            return True, history[number]
        return False, None
    def validate_page(self, text: str) -> dict:
        """验证单页内容"""
        info = self.extract_info(text)
        validation = {
            "状态": True,
        }
        # 添加额外验证逻辑（如日期格式校验）
        if "开票日期" in info and not self._validate_date(info["开票日期"]):
            validation["状态"] = False
        # 返回合并后的字典
        return {**info, **validation}

    def extract_info(self, text: str) -> Dict[str, str]:
        """Extract and clean invoice information"""
        info = {}
        for key, pattern in self.patterns.items():
            match = re.search(pattern, text)
            if match:
                # Clean spaces and take last 20 digits (compatible with OCR misrecognition)
                if key in ["发票号码"]:
                    cleaned = match.group(1).replace(" ", "")[-20:]
                else:
                    cleaned = match.group(1).replace(" ", "")
                info[key] = cleaned
        return info

    def _validate_date(self, date_str: str) -> bool:
        """验证日期格式"""
        try:
            datetime.datetime.strptime(date_str, "%Y年%m月%d日")
            return True
        except ValueError:
            return False
    def validate(self, info: Dict[str, str]) -> str:
        """Execute complete validation process"""
        errors = []

        # Required field check
        required_fields = ["发票号码"]
        for field in required_fields:
            if field not in info:
                errors.append(f"❌ 缺少必要字段：{field}")
                return "\n".join(errors)
        return "✅ 发票有效" if not errors else "\n".join(errors)


# -------------------- Image Processing Module (Optimized) --------------------
class ImageProcessor:
    """Image processing class"""

    @staticmethod
    def preprocess_image(image: np.ndarray, optimize_mode: int) -> np.ndarray:
        """Image preprocessing"""
        processed_image = image.copy()

        # Grayscale enhancement
        if optimize_mode in [1, 4, 6]:
            if len(processed_image.shape) == 3:  # Color image
                processed_image = cv2.cvtColor(processed_image, cv2.COLOR_BGR2GRAY)

        # Binarization
        if optimize_mode in [2, 4, 5, 6]:
            if len(processed_image.shape) == 3:  # Color image
                processed_image = cv2.cvtColor(processed_image, cv2.COLOR_BGR2GRAY)
            _, processed_image = cv2.threshold(processed_image, 128, 255, cv2.THRESH_BINARY)

        # Denoising
        if optimize_mode in [3, 5, 6]:
            processed_image = cv2.medianBlur(processed_image, 5)

        return processed_image


# -------------------- OCR Engine Module (Optimized) --------------------
class OCREngine:
    """OCR engine class"""

    def __init__(self, precision: str):
        self.lock = threading.Lock()
        self.max_retries = 3
        self.timeout = 30
        self.max_memory_usage = 0
        self.last_check_time = time.time()
        self.initialized = False
        self.precision = precision
        self._init_engine()

    def _init_engine(self):
        try:
            if not self.initialized:
                self.paddleocr = PaddleOCR(
                    use_angle_cls=True,
                    det_model_dir=resource_path("models/ch_PP-OCRv4_det_infer"),
                    rec_model_dir=resource_path("models/ch_PP-OCRv4_rec_infer"),
                    cls_model_dir=resource_path("models/ch_ppocr_mobile_v2.0_cls_infer"),
                    cpu_threads=4
                )
                self.initialized = True
        except Exception as e:
            logging.error(f"OCR引擎初始化失败: {str(e)}")
            if hasattr(self, 'root'):
                self.root.destroy()
    def _check_system_resources(self):
        """检查系统资源"""
        # 限制处理频率
        if time.time() - self.last_check_time < 0.5:
            time.sleep(0.1)
        self.last_check_time = time.time()
    def _parse_paddle_result(self, result):
        """增强版解析方法，兼容不同版本的PaddleOCR结果格式"""
        recognized_text = []

        try:
            # 处理单页结果
            if isinstance(result, (list, tuple)):
                for line in result:
                    if isinstance(line, (list, tuple)):
                        line_text = []
                        for word_info in line:
                            # 兼容不同版本的结果格式
                            if isinstance(word_info, (list, tuple)) and len(word_info) >= 2:
                                text_data = word_info[1]
                                if isinstance(text_data, (list, tuple)):
                                    text = str(text_data[0]) if len(text_data) > 0 else ""
                                else:
                                    text = str(text_data)
                                line_text.append(text)
                        recognized_text.append(' '.join(line_text))

            # 处理多页结果
            elif isinstance(result, dict) and 'data' in result:
                for item in result['data']:
                    if 'text' in item:
                        recognized_text.append(str(item['text']))

            return '\n'.join(recognized_text)

        except Exception as e:
            logging.error(f"解析PaddleOCR结果失败: {str(e)}")
            return f"OCR解析错误: {str(e)}"

    def recognize_sync(self, image_path, processed_image=None):
        with self.lock:
            try:
                logging.info(f"开始识别: {image_path}")
                # 每次识别前重新初始化 OCR 引擎（避免脏数据）
                if hasattr(self, 'paddleocr'):
                    del self.paddleocr

                # 重新初始化 PaddleOCR（限制显存）
                self.paddleocr = PaddleOCR(
                    rec_batch_num=1,  # 减小识别批量
                    det_batch_num=1  # 减小检测批量
                )

                # 执行识别
                result = self.paddleocr.ocr(image_path, cls=True)
                return self._parse_paddle_result(result)
            except Exception as e:
                logging.error(f"识别失败: {str(e)}", exc_info=True)
                raise
            finally:
                # 确保释放资源
                if hasattr(self, 'paddleocr'):
                    del self.paddleocr
                paddle.device.cuda.empty_cache()
    def recognize(self, image_path: str) -> str:
        """Perform OCR recognition"""
        try:
            return self._recognize_with_paddleocr(image_path)
        finally:
            # 确保释放资源
            if hasattr(self, 'paddleocr') and hasattr(self.paddleocr, 'close'):
                self.paddleocr.close()

    def _recognize_with_paddleocr(self, image_path: str) -> str:
        """Use PaddleOCR for OCR recognition"""
        result = self.paddleocr.ocr(image_path, cls=True)

        recognized_text = ""
        if isinstance(result, list):
            for line in result:
                if isinstance(line, list):
                    for word_info in line:
                        if isinstance(word_info, list) and len(word_info) > 1:
                            text, confidence = word_info[1]
                            recognized_text += text + ' '
                elif isinstance(line, list) and len(line) > 1:
                    text, confidence = line[1]
                    recognized_text += text + ' '

        return recognized_text.strip()


# -------------------- File Processing Module (Optimized) --------------------



class FileProcessor:
    """File processing class"""
    @staticmethod
    def sync_timeout(seconds):
        """同步超时装饰器"""
        def decorator(func):
            @functools.wraps(func)
            def wrapper(*args, **kwargs):
                result = None
                exception = None
                event = threading.Event()

                def target():
                    nonlocal result, exception
                    try:
                        result = func(*args, **kwargs)
                    except Exception as e:
                        exception = e
                    finally:
                        event.set()

                thread = threading.Thread(target=target, daemon=True)
                thread.start()

                if not event.wait(seconds):
                    raise TimeoutError(f"Operation timed out after {seconds} seconds")
                if exception:
                    raise exception
                return result
            return wrapper
        return decorator
    @staticmethod
    @staticmethod
    def extract_text_from_pdf(
            pdf_path: str,
            ocr_engine: OCREngine,
            image_processor: ImageProcessor,
            optimize_mode: int,
            cpu_threshold: int,
            progress_callback: callable = None,
            result_callback: callable = None
    ) -> List[Dict]:
        results = []
        temp_dir = tempfile.mkdtemp(prefix="pdf_ocr_")
        cpu_monitor = CPUMonitor(threshold=cpu_threshold)
        start_time = time.time()

        try:
            with fitz.open(pdf_path) as pdf_doc:
                total_pages = len(pdf_doc)
                logging.info(f"开始处理PDF: {pdf_path} 总页数: {total_pages}")
                for page_num in range(total_pages):
                    cpu_monitor.adjust_speed()
                    throttle_time = cpu_monitor.get_throttle_time()
                    time.sleep(throttle_time)

                    temp_image = os.path.join(
                        temp_dir,
                        f"page_{page_num}_{os.urandom(4).hex()}.png"
                    )

                    try:
                        page = pdf_doc.load_page(page_num)
                        pix = page.get_pixmap(dpi=300)
                        pix.save(temp_image)

                        image = cv2.imread(temp_image)
                        if image is None:
                            raise ValueError(f"无法加载图像: {temp_image}")

                        processed_image = image_processor.preprocess_image(image, optimize_mode)

                        # 确保 self.ocr_engine 是 OCREngine 的实例
                        if not isinstance(ocr_engine, OCREngine):
                            raise TypeError("ocr_engine 必须是 OCREngine 的实例")

                        page_text = ocr_engine.recognize_sync(temp_image, processed_image)
                        result = {
                            "页码": page_num + 1,
                            "文本": page_text,
                            "处理时间": {
                                "总耗时": f"{(time.time() - start_time):.2f}s",
                            },
                            "状态": "成功"
                        }

                    except Exception as e:
                        logging.error(f"页面 {page_num + 1} 处理失败", exc_info=True)
                        result = {
                            "页码": page_num + 1,
                            "文本": f"处理错误: {str(e)}",
                            "处理时间": None,
                            "状态": "失败"
                        }
                    results.append(result)

                    if result_callback:
                        result_callback(result, page_num + 1, total_pages)

                    if progress_callback:
                        progress_callback(page_num + 1, total_pages)

                success_count = sum(1 for r in results if r["状态"] == "成功")
                logging.info(
                    f"PDF处理完成: {success_count}/{total_pages} 页成功, "
                    f"总耗时: {time.time() - start_time:.2f}秒"
                )

                return results

        except Exception as e:
            logging.critical(f"PDF处理严重错误: {str(e)}", exc_info=True)
            raise RuntimeError(f"PDF处理失败: {str(e)}") from e

        finally:
            max_retries = 3
            for attempt in range(max_retries):
                try:
                    shutil.rmtree(temp_dir, ignore_errors=True)
                    if not os.path.exists(temp_dir):
                        break
                    time.sleep(0.5 * (attempt + 1))
                except Exception as e:
                    logging.error(
                        f"临时目录清理失败（尝试 {attempt + 1}/{max_retries}）: {str(e)}"
                    )
                    if attempt == max_retries - 1:
                        logging.warning(f"无法完全清理临时目录: {temp_dir}")

    @sync_timeout(seconds=25)  # 直接装饰方法
    def ocr_task(self, temp_image, processed_image):
        return self.ocr_engine.recognize_sync(temp_image, processed_image)

    @staticmethod
    def extract_text_from_image(image_path: str, ocr_engine: OCREngine, image_processor: ImageProcessor,
                                optimize_mode: int) -> str:
        """Extract text from image"""
        # Read image
        image = cv2.imread(image_path)
        if image is None:
            raise ValueError(f"Cannot load image: {image_path}")

        # Image preprocessing
        processed_image = image_processor.preprocess_image(image, optimize_mode)

        # OCR recognition
        return ocr_engine.recognize(image_path, processed_image)

    @staticmethod
    def write_to_excel(data: str, output_path: str, row: int = 1, col: int = 1) -> None:
        """Write data to Excel"""
        workbook = Workbook()
        sheet = workbook.active
        sheet.cell(row=row, column=col, value=data)
        workbook.save(output_path)


# Main program
class ClassMain:
    def __init__(self, root):
        self.root = root
        self.root.geometry('1200x800')
        self.current_language = "zh_CN"  # Default language
        self.selected_language = "zh_CN"  # Language selected but not yet applied
        self.config_file = "config.ini"
        self.history = {}
        self.config = configparser.ConfigParser()
        self.check_duplicate_var = tk.BooleanVar()
        self.check_duplicate_var = tk.BooleanVar()
        self.invoice_recognition_var = tk.BooleanVar()
        self.warning_prompt_var = tk.BooleanVar()
        self.retain_text_var = tk.BooleanVar(value=False)
        self.precision_var = tk.StringVar(value="fp32")
        self.cpu_threshold_var = tk.IntVar(value=65)
        self.font_family_var = tk.StringVar()
        self.font_size_var = tk.IntVar()
        self.model_dir = resource_path('models')
        self.style = ttk.Style(theme="minty")  # 可选主题: minty, darkly, solar, superhero
        self.style.configure("TButton", font=('微软雅黑', 10), padding=6)
        self.style.map("TButton",
                       background=[('active', '!disabled', '#4CAF50'), ('pressed', '#45a049')]
                       )
        # 初始化默认值
        self.hw_support = {
            "gpu": False,
            "mkldnn": False,
            "tensorrt": False,
            "fp16": False,
            "bf16": False,
            "gpu_info": "未检测到"
        }
        # 更新界面状态
        self.load_config()
        self.root.title(languages[self.current_language]["title"])
        try:
            from ctypes import windll
            windll.shcore.SetProcessDpiAwareness(1)
        except:
            pass
        # Initialize processors
        self.image_processor = ImageProcessor()
        self.ocr_engine = None  # Lazy initialization
        self.canvas_image_id = None  # 初始化图像ID
        self.file_processor = FileProcessor()
        self.crop_area = (0, 0, 0, 0)
        self._preview_job = None
        self.current_mode = "pan"  # 默认平移模式
        # 添加缩放相关变量
        self.zoom_scale = 1.0  # 当前缩放比例
        self.min_zoom = 0.1  # 最小缩放比例
        self.max_zoom = 10.0  # 最大缩放比例
        self.image_position = (0, 0)  # 图片当前位置
        # 修改初始化语言设置
        self.current_language = self.config.get("DEFAULT", "language", fallback="zh_CN")
        self.selected_language = self.current_language
        # Initialize lists with translated values
        self.optimize_list = [
            languages[self.current_language]["optimize_none"],
            languages[self.current_language]["optimize_grayscale"],
            languages[self.current_language]["optimize_binary"],
            languages[self.current_language]["optimize_denoise"],
            languages[self.current_language]["optimize_grayscale_binary"],
            languages[self.current_language]["optimize_binary_denoise"],
            languages[self.current_language]["optimize_all"]
        ]

        self.ocr_module_list = [languages[self.current_language]["ocr_paddle"]]

        self.optimization_methods = [
            languages[self.current_language]["denoise_strength"],
            languages[self.current_language]["sharpen_strength"],
            languages[self.current_language]["brightness"],
            languages[self.current_language]["contrast"],
            languages[self.current_language]["edge_enhancement"]
        ]
        self.optimization_parameters = [
            languages[self.current_language]["denoise_strength"],
            languages[self.current_language]["sharpen_strength"],
            languages[self.current_language]["brightness"],
            languages[self.current_language]["contrast"],
            languages[self.current_language]["edge_enhancement"]
        ]
        # Set window icon
        icon_path = resource_path('t1.ico')
        if os.path.exists(icon_path):
            try:
                icon_image = Image.open(icon_path)
                icon_photo = ImageTk.PhotoImage(icon_image)
                self.root.iconphoto(True, icon_photo)
            except Exception as e:
                print(f"Icon loading failed: {str(e)}")
        else:
            print(f"Icon file {icon_path} not found")
        self.selected_conversion_option = None
        self.conversion_output = None
        self.create_widgets()
        self.extracted_info = {}
        self.validation_result = ""
        self.preview_image = None
        # 创建加载动画组件
        self.create_loading_animation()
        self.init_canvas_moveto()
        self.update_hardware_status_labels()
        print(time.time() - nowtimer)
        self.root.protocol("WM_DELETE_WINDOW", self.on_close)

    # 更新状态显示方法
    def update_hardware_status_labels(self):
        """更新硬件状态显示"""
        try:
            # 计算精度
            precisions = ["FP32"]
            if self.hw_support["fp16"]:
                precisions.append("FP16")
            if self.hw_support["bf16"]:
                precisions.append("BF16")
            self.precision_status.config(text=f"计算精度支持: {'/'.join(precisions)}")

        except Exception as e:
            logging.error(f"更新状态显示失败: {str(e)}")
            self.precision_status.config(text="计算精度: 显示错误")

    def load_history_from_config(self):
        self.history = {}
        if self.config.has_section("HISTORY"):
            for key, value in self.config.items("HISTORY"):
                if key.isdigit() or key.isdigit():
                    self.history[key] = value

    def start_health_check(self):
        """每5分钟检查一次引擎状态"""

        def check():
            if not self.ocr_engine.initialized:
                try:
                    self.ocr_engine._init_engine()
                except:
                    logging.error("引擎健康检查失败")
            self.root.after(300000, check)  # 5分钟间隔

        self.root.after(60000, check)  # 1分钟后启动
    def init_ocr_engine(self):
        """增强的引擎初始化"""
        try:
            if not hasattr(self, 'ocr_engine') or not self.ocr_engine:
                self.ocr_engine = OCREngine(precision=self.precision_var.get())
        except Exception as e:
            messagebox.showerror("致命错误", f"OCR引擎初始化失败: {str(e)}")
            self.root.destroy()
    # 使用 Progressbar 实现加载动画
    def setup_canvas_events(self):
        # 绑定缩放事件（跨平台兼容）
        self.canvas.bind("<MouseWheel>", self.zoom_image)  # Windows/Mac
        self.canvas.bind("<Button-4>", self.zoom_image)  # Linux向上
        self.canvas.bind("<Button-5>", self.zoom_image)  # Linux向下
    def start_loading(self):
        self.progress.start()
        self.progress.configure(mode='indeterminate', bootstyle="striped-info")
        self.root.after(3000, self.stop_loading)

    def stop_loading(self):
        self.progress.stop()
        self.progress.configure(mode='determinate')
    def sync_image_path(self, path):
        # 更新识别选项卡
        self.file_path.config(state='normal')
        self.file_path.delete(0, tk.END)
        self.file_path.insert(0, path)
        self.file_path.config(state='disabled')

        # 更新优化选项卡
        self.optimize_file_path.config(state='normal')
        self.optimize_file_path.delete(0, tk.END)
        self.optimize_file_path.insert(0, path)
        self.optimize_file_path.config(state='disabled')

        # 触发预览更新
        self.update_preview()
        self.load_optimize_preview(path)

    def on_mouse_down(self, event):
        """统一鼠标按下处理"""
        if self.current_mode == "crop":
            self.on_press(event)
        elif self.current_mode == "pan":
            self.on_pan_start(event)

    def on_mouse_drag(self, event):
        """统一鼠标拖动处理"""
        if self.current_mode == "crop":
            self.on_drag(event)
        elif self.current_mode == "pan":
            self.pan_image(event)

    def on_mouse_up(self, event):
        """统一鼠标释放处理"""
        if self.current_mode == "crop":
            self.on_release(event)
            self.switch_to_pan_mode()  # 裁剪完成后自动切回平移模式
        elif self.current_mode == "pan":
            self.on_pan_end(event)
    def switch_to_pan_mode(self):
        """切换回平移模式"""
        self.current_mode = "pan"
        self.mode_label.config(
            text=f"{languages[self.current_language]['current_mode']}: "f"{languages[self.current_language]['pan_mode' if self.current_mode == 'pan' else 'crop_mode']}")
        self.crop_enabled = False
    def start_crop_selection(self):
        """开始裁剪选择"""
        self.crop_enabled = True
        self.current_mode = "crop"
        self.mode_label.config(text=f"{languages[self.current_language]['current_mode']}: "f"{languages[self.current_language]['pan_mode' if self.current_mode == 'pan' else 'crop_mode']}")
        self.canvas.delete("all")
        self.show_edit_image(self.original_edit_image)
        # 显示用户提示
        messagebox.showinfo(languages[self.current_language]["info"],
                            languages[self.current_language]["crop_tips"])
        self.crop_area = (0, 0, 0, 0)
        self.crop_coords.set("")
    def on_press(self, event):
        """鼠标按下事件"""
        self.canvas.delete("all")
        self.show_edit_image(self.original_edit_image)
        if self.crop_enabled:
            self.start_x = event.x
            self.start_y = event.y
            self.rect = self.canvas.create_rectangle(self.start_x, self.start_y, self.start_x, self.start_y,
                                                     outline='red', width=2)

    def on_drag(self, event):
        """鼠标拖动事件"""
        if self.rect and self.crop_enabled:
            self.canvas.coords(self.rect, self.start_x, self.start_y, event.x, event.y)
            # 实时预览裁剪区域
            self.update_crop_preview()

    def on_release(self, event):
        """鼠标释放事件"""
        if self.crop_enabled:
            x1, y1 = self.start_x, self.start_y
            x2, y2 = event.x, event.y
            # 边界检查
            x1, y1, x2, y2 = self.check_crop_boundaries(x1, y1, x2, y2)
            self.crop_coords.set(f"({x1},{y1}) - ({x2},{y2})")
            self.crop_area = (x1, y1, x2, y2)

    def check_crop_boundaries(self, x1, y1, x2, y2):
        """确保裁剪区域在有效范围内"""
        canvas_w = self.canvas.img_display_width
        canvas_h = self.canvas.img_display_height

        # 约束坐标在 Canvas 范围内
        x1 = max(0, min(x1, canvas_w))
        y1 = max(0, min(y1, canvas_h))
        x2 = max(0, min(x2, canvas_w))
        y2 = max(0, min(y2, canvas_h))

        # 确保 x2 > x1 且 y2 > y1
        if x2 < x1:
            x1, x2 = x2, x1
        if y2 < y1:
            y1, y2 = y2, y1

        return x1, y1, x2, y2
    def update_crop_preview(self):
        """更新裁剪区域预览"""
        if self.start_x is None or self.start_y is None:
            return
        x1, y1 = self.start_x, self.start_y
        x2, y2 = self.canvas.canvasx(self.canvas.winfo_pointerx()), self.canvas.canvasy(self.canvas.winfo_pointery())
        # 绘制半透明遮罩
        self.canvas.delete("preview")
        self.canvas.create_rectangle(0, 0, self.canvas.winfo_width(), self.canvas.winfo_height(), fill='gray',
                                     stipple='gray50', tags="preview")
        self.canvas.create_rectangle(x1, y1, x2, y2, fill='', stipple='', tags="preview")

    def update_progress(self, current, total):
        # 添加双重缓冲
        if not hasattr(self, '_last_progress'):
            self._last_progress = 0

        if abs(current - self._last_progress) >= max(1, total // 20):
            self.progress["value"] = int((current / total) * 100)
            self.progress_label.config(text=f"{int((current / total) * 100)}%")
            self.count_label.config(text=f"{current}/{total}")
            self._last_progress = current
            self.root.update_idletasks()
    def update_language(self):
        """Update UI elements with the current language"""
        # Update window title
        self.root.title(languages[self.current_language]["title"])

        for i, tab_name in enumerate(["tab_edit","tab_recognize", "tab_convert", "tab_optimize", "tab_settings","tab_about"]):
            self.notebook.tab(i, text=languages[self.current_language][tab_name])

        # Define UI elements and their corresponding text keys
        ui_elements = {
            # Recognize tab
            "file_path_label": "file_path",
            "browse_button": "browse",
            "settings_frame": "settings",
            "excel_name_label": "excel_name",
            "optimize_mode_label": "recognition_optimize",
            "ocr_module_label": "ocr_module",
            "start_button": "start_recognition",
            "export_button": "export_to_a1",
            "row_label": "row",
            "col_label": "column",
            "export_pos_button": "export_to_position",
            "result_frame": "recognition_result",
            "copy_button": "copy_content",
            "clear_button": "clear_content",
            "info_frame": "invoice_key_info",
            # Optimize tab
            "optimize_file_label": "select_image",
            "optimize_browse_button": "browse",
            "options_frame": "optimization_options",
            "save_button": "save_result",
            "reset_button": "reset",
            "before_preview_frame": "original_preview",
            "after_preview_frame": "processed_preview",
            "optimize_settings_frame": "settings",
            # Convert tab
            "conversion_label": "select_conversion",
            "display_frame": "conversion_result",
            "output_image_format_label": "output_format",
            "merge_multiple_label": "merge_multiple_images",
            # About tab
            "developer_label": "developer",
            "dev_date_label": "dev_date",
            "version_label": "current_version",
            "update_button": "check_update",
            "language_label": "language",
            "apply_button": "apply_changes",
            "log_frame": "update_log",
            "copyright_label": "copyright",

            "edit_file_label": "edit_select_image",
            "browse_edit_btn": "browse",
            "resize_frame": "resize_settings",
            "width_label": "width",
            "height_label": "height",
            "ratio_check": "keep_ratio",
            "aspect_ratio": "aspect_ratio",
            "image_save_btn": "save_edit",
            "crop_label": "crop_coords",
            "crop_btn": "start_crop",
            "crop_frame": "crop_settings",
            "duplicate_frame": "invoice_duplicate_check",
            "check_duplicate": "invoice_duplicate_check",
            "save_btn": "save_settings",
            "language_frame": "language",
            "clear_selected_btn": "clear_selected",
            "clear_all_btn": "clear_all",
            "check_warning": "warning_prompt",
            # Conversion options
            "conversion_options": [
                "conversion_all",
                "conversion_pdf_to_image",
                "conversion_image_to_pdf",
                "conversion_txt_to_pdf",
                "conversion_merge_pdf",
                "conversion_info_only"
            ],

            # Optimize list
            "optimize_mode": [
                "optimize_none",
                "optimize_grayscale",
                "optimize_binary",
                "optimize_denoise",
                "optimize_grayscale_binary",
                "optimize_binary_denoise",
                "optimize_all"
            ],

            # OCR module list
            "ocr_module": ["ocr_paddle"],

            # Optimization parameters
            "optimization_parameters": [
                "denoise_strength",
                "sharpen_strength",
                "brightness",
                "contrast",
                "edge_enhancement"
            ]
        }

        # Update UI elements based on the defined mappings
        for element, text_key in ui_elements.items():
            if hasattr(self, element):
                if element == "result_tree_columns":
                    for col, key in text_key.items():
                        self.result_tree.heading(col, text=languages[self.current_language][key])
                elif element == "conversion_options":
                    self.conversion_options['values'] = [languages[self.current_language][key] for key in text_key]
                    self.conversion_options.current(0)
                elif element == "optimize_mode":
                    current_value = self.optimize_mode.get()
                    self.optimize_list = [languages[self.current_language][key] for key in text_key]
                    self.optimize_mode['values'] = self.optimize_list
                    try:
                        index = self.optimize_list.index(current_value)
                        self.optimize_mode.current(index)
                    except ValueError:
                        self.optimize_mode.current(0)
                elif element == "ocr_module":
                    current_value = self.ocr_module.get()
                    self.ocr_module_list = [languages[self.current_language][key] for key in text_key]
                    self.ocr_module['values'] = self.ocr_module_list
                    try:
                        index = self.ocr_module_list.index(current_value)
                        self.ocr_module.current(index)
                    except ValueError:
                        self.ocr_module.current(0)
                elif element == "optimization_parameters":
                    # Update optimization parameter labels
                    for param in text_key:
                        label1 = getattr(self, f"{languages['zh_CN'][param]}_label", None)
                        label2 = getattr(self, f"{languages['en_US'][param]}_label", None)
                        if label1:
                            label1.config(text=languages[self.current_language][param])
                        elif label2:
                            label2.config(text=languages[self.current_language][param])
                elif element == "version_label":
                    getattr(self, element).config(text=languages[self.current_language][text_key].format(version=latest_version))
                else:
                    widget = getattr(self, element)
                    if isinstance(widget, (tk.Label, ttk.Label, ttk.LabelFrame, tk.Button,ttk.Button)):
                        widget.config(text=languages[self.current_language][text_key])
                    elif isinstance(widget, ttk.Checkbutton):
                        widget.config(text=languages[self.current_language][text_key])
        # Update preset ratios in edit tab
        if hasattr(self, 'ratio_combobox'):
            current_value = self.ratio_var.get()
            new_values = list(languages[self.current_language]["preset_ratios"].values())
            self.ratio_combobox['values'] = new_values
            try:
                index = new_values.index(current_value)
                self.ratio_combobox.current(index)
            except ValueError:
                self.ratio_combobox.current(0)
        # Update log text
        if hasattr(self, 'logtext'):
            self.logtext.config(state='normal')
            self.logtext.delete(1.0, tk.END)
            log = ch_log if self.current_language == 'zh_CN' else en_log
            for log_entry in log:
                self.logtext.insert(tk.END, log_entry + '\n\n')
            self.logtext.config(state='disabled')

    def clear_selected_history(self):
        """清除当前显示的发票记录"""
        # 获取当前发票信息
        number = self.extracted_info.get("发票号码", "")

        if not number:
            messagebox.showwarning(
                languages[self.current_language]["warning"],
                languages[self.current_language]["no_records_selected"]
            )
            return

        history_key = number

        # 检查内存中的记录
        if history_key in self.history:
            # 从内存删除
            del self.history[history_key]

            # 更新配置文件
            if self.config.has_section("HISTORY") and self.config.has_option("HISTORY", number):
                self.config.remove_option("HISTORY", number)
                self._save_config_impl()

            messagebox.showinfo(
                languages[self.current_language]["success"],
                languages[self.current_language]["selected_cleared"]
            )
        else:
            messagebox.showinfo(
                languages[self.current_language]["info"],
                languages[self.current_language]["no_records_selected"]
            )

    def clear_all_history(self):
        """清除所有历史记录"""
        total_count = len(self.history)
        self.history.clear()

        # 更新配置文件
        if self.config.has_section("HISTORY"):
            self.config.remove_section("HISTORY")
        self.config.add_section("HISTORY")
        self._save_config_impl()

        # 更新UI
        self.result_tree.delete(*self.result_tree.get_children())
        messagebox.showinfo(
            languages[self.current_language]["success"],
            languages[self.current_language]["history_cleared"].format(count=total_count)
        )

    def _purge_config_history(self):
        """专用历史记录清理方法"""
        if self.config.has_section("HISTORY"):
            self.config.remove_section("HISTORY")
        self.config.add_section("HISTORY")
        self._save_config_impl()  # 确保写入空配置
    def _sync_config(self):
        """专用配置同步方法"""
        if self.config.has_section("HISTORY"):
            self.config.remove_section("HISTORY")
        self.config.add_section("HISTORY")

        # 写入当前内存状态
        for number, timestamp in self.history.items():
            self.config.set("HISTORY", number, timestamp)

        self._save_config_impl()

    def _save_config_impl(self):
        """配置保存底层实现"""
        with open(self.config_file, "w") as configfile:
            self.config.write(configfile)
    def create_ratio_controls(self, parent_frame):
        """创建比例控制组件"""
        # 比例选择框
        self.ratio_frame = ttk.Frame(parent_frame)
        self.ratio_frame.pack(fill=tk.X, pady=5)

        self.aspect_ratio = ttk.Label(self.ratio_frame, text=languages[self.current_language]["aspect_ratio"])
        self.aspect_ratio.pack(side=tk.LEFT)

        self.ratio_var = tk.StringVar()
        self.ratio_combobox = ttk.Combobox(
            self.ratio_frame,
            textvariable=self.ratio_var,
            values=list(languages[self.current_language]["preset_ratios"].values()),
            state="readonly"
        )
        self.ratio_combobox.pack(side=tk.LEFT, padx=5)
        self.ratio_combobox.bind("<<ComboboxSelected>>", self.on_ratio_change)
        self.ratio_combobox.current(0)

        # 当前比例存储
        self.current_ratio = None  # (width, height) 元组格式
        self.custom_ratio = None

    def on_ratio_change(self, event=None):
        """处理比例选择变化"""
        selected = self.ratio_var.get()
        preset_map = {
            v: k for k, v in languages[self.current_language]["preset_ratios"].items()
        }

        # 处理自定义比例
        if preset_map[selected] == "custom":
            ratio_str = simpledialog.askstring(
                languages[self.current_language]["custom_ratio_prompt"],
                languages[self.current_language]["custom_ratio_prompt"]
            )
            if ratio_str:
                try:
                    parts = ratio_str.split(":")
                    if len(parts) != 2:
                        raise ValueError
                    w, h = map(float, parts)
                    self.custom_ratio = (w, h)
                    self.current_ratio = self.custom_ratio
                except:
                    messagebox.showerror(
                        languages[self.current_language]["error"],
                        languages[self.current_language]["invalid_ratio_format"]
                    )
                    self.ratio_combobox.current(0)
                    self.current_ratio = self.original_ratio
            else:
                self.ratio_combobox.current(0)
                self.current_ratio = self.original_ratio
        else:
            # 处理预设比例
            ratio_map = {
                "original": self.original_ratio,
                "3:2": (3, 2),
                "4:3": (4, 3),
                "16:9": (16, 9),
                "1:1": (1, 1)
            }
            self.current_ratio = ratio_map[preset_map[selected]]

        # 如果保持比例开启，自动调整尺寸
        if self.keep_ratio.get():
            self.apply_aspect_ratio()

    def apply_aspect_ratio(self, changed_field=None):
        """应用宽高比例规则"""
        if not self.current_ratio or not self.keep_ratio.get():
            return

        try:
            # 获取当前有效值
            width_str = self.width_var.get().strip()
            height_str = self.height_var.get().strip()

            # 确定哪个字段被修改
            if changed_field == "width" and width_str:
                new_width = float(width_str)
                new_height = new_width * (self.current_ratio[1] / self.current_ratio[0])
                self.height_var.set(str(round(new_height)))
            elif changed_field == "height" and height_str:
                new_height = float(height_str)
                new_width = new_height * (self.current_ratio[0] / self.current_ratio[1])
                self.width_var.set(str(round(new_width)))
            elif width_str and height_str:  # 两个值都存在时优先宽度
                new_width = float(width_str)
                new_height = new_width * (self.current_ratio[1] / self.current_ratio[0])
                self.height_var.set(str(round(new_height)))
        except ValueError:
            pass

    def batch_check_pdf(self, pages):
        """批量检查PDF多页"""
        results = []
        validator = InvoiceValidator()
        for page_data in pages:
            validation = validator.validate_page(page_data["文本"])
            result = {
                "页码": page_data["页码"],
                "发票号码": validation.get("发票号码", "无"),
                "开票日期": validation.get("开票日期", "无"),
                "状态": "有效发票" if validation["状态"] else "非发票",
            }
            results.append(result)

        self.update_results_table(results)
    def show_crop_tips(self):
        """显示裁剪操作提示"""
        tips = [
            "1. 在图片上点击并拖动以选择裁剪区域",
            "2. 松开鼠标后会自动调整裁剪区域到有效范围",
            "3. 裁剪区域会实时预览",
            "4. 点击'保存修改'按钮保存裁剪后的图片"
        ]
        messagebox.showinfo(languages[self.current_language]["info"], "\n".join(tips))

    def change_theme(self):
        theme_map = {
            languages[self.current_language]["style_minty"]: "minty",
            languages[self.current_language]["style_darkly"]: "darkly",
            languages[self.current_language]["style_solar"]: "solar",
            languages[self.current_language]["style_superhero"]: "superhero"
        }

        selected_theme = theme_map[self.theme_combobox.get()]

        # 更新样式
        if selected_theme in ttk.Style().theme_names():  # 检查主题是否有效
            self.style.theme_use(selected_theme)
        else:
            messagebox.showerror(languages[self.current_language]["error"],
                                 f"无效的主题：{selected_theme}")
            return

        # 更新配置
        self.config.set("DEFAULT", "theme", selected_theme)
        self.save_config(False)

        # 刷新界面
        self.update_language()
    def create_settings_tab(self, tab):
        def on_enter(e):
            e.widget.config(bootstyle=SUCCESS)

        def on_leave(e):
            e.widget.config(bootstyle=(SUCCESS,OUTLINE))
        """创建设置选项卡"""
        main_frame = ttk.Frame(tab)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        style = ttk.Style()
        style.configure("Customred.TLabel", foreground="red", background="white")
        style.configure("Customblue.TLabel", foreground="blue", background="white")
        # 语言设置
        self.language_frame = ttk.LabelFrame(main_frame, text=languages[self.current_language]["language"])
        self.language_frame.pack(fill=tk.X, pady=5, padx=5)

        self.language_combobox = ttk.Combobox(
            self.language_frame,
            values=["简体中文", "English"],
            state="readonly"
        )
        self.language_combobox.current(0 if self.current_language == "zh_CN" else 1)
        self.language_combobox.pack(pady=5)
        # 在语言设置之后添加字体设置
        font_frame = ttk.LabelFrame(main_frame, text="字体设置")
        font_frame.pack(fill=tk.X, pady=5, padx=5)

        # 字体选择
        ttk.Label(font_frame, text="字体:").pack(side=tk.LEFT)
        self.font_family = ttk.Combobox(font_frame, textvariable=self.font_family_var)
        self.font_family['values'] = list(tk.font.families())
        self.font_family.pack(side=tk.LEFT, padx=5)

        # 字号选择
        ttk.Label(font_frame, text="字号:").pack(side=tk.LEFT)
        self.font_size = ttk.Spinbox(font_frame, from_=8, to=24, textvariable=self.font_size_var, width=5)
        self.font_size.pack(side=tk.LEFT, padx=5)
        # 主题切换
        self.theme_frame = ttk.LabelFrame(main_frame, text=languages[self.current_language]["theme_switch"])
        self.theme_frame.pack(fill=tk.X, pady=5, padx=5)

        # 获取可用主题列表
        self.themes = {
            "minty": languages[self.current_language]["style_minty"],
            "darkly": languages[self.current_language]["style_darkly"],
            "solar": languages[self.current_language]["style_solar"],
            "superhero": languages[self.current_language]["style_superhero"]
        }

        self.theme_combobox = ttk.Combobox(
            self.theme_frame,
            values=list(self.themes.values()),
            state="readonly"
        )
        self.theme_combobox.pack(pady=5)

        # 设置当前主题
        current_theme = self.config.get("DEFAULT", "theme", fallback="minty")
        current_theme_name = self.themes.get(current_theme, languages[self.current_language]["style_minty"])
        self.theme_combobox.set(current_theme_name)

        # 绑定主题切换事件
        self.theme_combobox.bind("<<ComboboxSelected>>", self.change_theme)
        # 识别设置
        self.identify_frame = ttk.LabelFrame(main_frame, text=languages[self.current_language]["identify_recognition"])
        self.identify_frame.pack(fill=tk.X, pady=5, padx=5)
        # 发票重复识别设置
        self.duplicate_frame = ttk.LabelFrame(self.identify_frame, text=languages[self.current_language]["invoice_duplicate_check"])
        self.duplicate_frame.pack(fill=tk.X, pady=5, padx=5)

        # 发票识别开关
        self.invoice_check = ttk.Checkbutton(
            self.duplicate_frame,
            text=languages[self.current_language]["enable_invoice_recognition"],
            variable=self.invoice_recognition_var
        )

        self.invoice_check.pack(pady=5)
        self.check_duplicate = ttk.Checkbutton(
            self.duplicate_frame,
            text=languages[self.current_language]["invoice_duplicate_check"],
            variable=self.check_duplicate_var
        )
        self.check_duplicate.pack(pady=5)
        self.check_duplicate_var.set(self.config.getboolean("DEFAULT", "check_duplicate", fallback=False))
        self.check_warning = ttk.Checkbutton(
            self.duplicate_frame,
            text=languages[self.current_language]["warning_prompt"],  # 需要添加到语言字典
            variable=self.warning_prompt_var
        )
        self.check_warning.pack(pady=5)

        self.clear_all_btn = ttk.Button(
            self.duplicate_frame,
            text=languages[self.current_language]["clear_all"],
            command=self.clear_all_history
        )
        self.clear_all_btn.pack(side=tk.TOP)
        # 新增加速选项框架
        acceleration_frame = ttk.LabelFrame(self.identify_frame, text=languages[self.current_language]["Identify_acceleration_options"])
        acceleration_frame.pack(fill=tk.X, pady=5, padx=5)
        cpu_frame = ttk.LabelFrame(self.identify_frame, text=languages[self.current_language]["CPU_throttling_settings"])
        cpu_frame.pack(fill=tk.X, pady=5, padx=5)

        ttk.Label(cpu_frame, text=languages[self.current_language]["CPU_max%use_frame"]).pack(side=tk.LEFT)
        ttk.Spinbox(cpu_frame,
                    from_=30, to=90,
                    increment=5,
                    textvariable=self.cpu_threshold_var,
                    width=5).pack(side=tk.LEFT, padx=5)

        ttk.Label(cpu_frame,
                  text=languages[self.current_language]["CPU_tips_frame"]).pack(side=tk.LEFT, pady=5)
        # 精度选择
        precision_frame = ttk.Frame(acceleration_frame)
        precision_frame.pack(fill=tk.X, pady=5)
        ttk.Label(precision_frame, text=languages[self.current_language]["computer_into"]).pack(side=tk.LEFT)
        ttk.Radiobutton(precision_frame, text="FP32", variable=self.precision_var,
                        value="fp32").pack(side=tk.LEFT)
        ttk.Radiobutton(precision_frame, text="FP16", variable=self.precision_var,
                        value="fp16").pack(side=tk.LEFT)
        # 计算精度状态
        self.precision_status = ttk.Label(precision_frame, text=languages[self.current_language]["computer_max_into"])
        self.precision_status.pack(anchor=tk.W, pady=5)
        ttk.Label(main_frame, text=languages[self.current_language]["setting_tips"], style="Customblue.TLabel").pack(
            fill=tk.X, pady=5, padx=5)
        ttk.Label(main_frame, text=languages[self.current_language]["setting_tips(d)"], style="Customred.TLabel").pack(
            fill=tk.X, pady=5, padx=5)
        # 保存按钮
        self.save_btn = ttk.Button(
            main_frame,
            text=languages[self.current_language]["save_settings"],
            command=self.save_config,
            bootstyle=(SUCCESS, OUTLINE)
        )
        self.save_btn.pack(pady=10)

        self.save_btn.bind("<Enter>", on_enter)
        self.save_btn.bind("<Leave>", on_leave)
    def load_all_setting(self):
        # 加载其他设置
        self.current_language = self.config.get("DEFAULT", "language", fallback="zh_CN")
        self.check_duplicate_var.set(self.config.getboolean("DEFAULT", "check_duplicate", fallback=False))
        self.warning_prompt_var.set(self.config.getboolean("DEFAULT", "show_warning", fallback=False))
        self.invoice_recognition_var.set(self.config.getboolean("DEFAULT", "invoice_recognition", fallback=False))
        self.cpu_threshold_var.set(self.config.getint("DEFAULT", "cpu_threshold", fallback=65))
        self.precision_var.set(self.config.get("DEFAULT", "precision", fallback="fp32"))
        self.font_family_var.set(self.config.get("DEFAULT", "font_family", fallback="微软雅黑"))
        self.font_size_var.set(self.config.getint("DEFAULT", "font_size", fallback=10))
    # 更新硬件状态标签
    def load_config(self):
        """加载配置文件"""
        if os.path.exists(self.config_file):
            self.config.read(self.config_file)

            # 正确加载历史记录
            self.history ={}
            if self.config.has_section("HISTORY"):
                for key, value in self.config.items("HISTORY"):
                    # 确保只加载有效的发票历史记录（号码:时间戳）
                    if (len(key) == 8 and key.isdigit()) or (len(key) == 20 and key.isdigit()):  # 假设发票号码是8/20位数字
                        self.history[key] = value

            self.load_all_setting()
            self.apply_font_settings()
        else:
            # 初始化默认配置
            self._init_default_config()

    def save_config(self, show_message=True):
        """保存配置文件"""
        # 清理旧配置
        for section in list(self.config.sections()):
            self.config.remove_section(section)
        # 保存DEFAULT设置
        self.config["DEFAULT"] = {
            "language": self.current_language,
            "check_duplicate": str(self.check_duplicate_var.get()),
            "show_warning": str(self.warning_prompt_var.get()),
            "theme": self.style.theme_use(),
            "invoice_recognition": str(self.invoice_recognition_var.get()),
            "cpu_threshold": str(self.cpu_threshold_var.get()),
            "precision": self.precision_var.get(),
            "font_family" : self.font_family_var.get(),
            "font_size" : self.font_size_var.get()
        }

        # 验证计算精度
        if self.precision_var.get() == "fp16" and not self.hw_support["fp16"]:
            messagebox.showwarning("警告", "当前硬件不支持FP16精度，已重置为FP32")
            self.precision_var.set("fp32")
        # 正确保存历史记录
        self.config.add_section("HISTORY")
        for number, timestamp in self.history.items():
            self.config.set("HISTORY", number, timestamp)

        # 写入文件
        with open(self.config_file, "w") as configfile:
            self.config.write(configfile)
        #super().save_config(show_message)
        if show_message:
            messagebox.showinfo(
                languages[self.current_language]["success"],
                languages[self.current_language]["save_settings"]
            )
    def apply_font_settings(self):
        font_style = (self.font_family_var.get(), self.font_size_var.get())
        self.style.configure(".", font=font_style)

        # 更新所有子组件字体
        self.update_font_recursive(self.root, font_style)

    def update_font_recursive(self, widget, font):
        try:
            widget.config(font=font)
        except tk.TclError:
            pass
        for child in widget.winfo_children():
            self.update_font_recursive(child, font)
    def _init_default_config(self):
        """初始化默认配置"""
        self.config["DEFAULT"] = {
            "language": "zh_CN",
            "check_duplicate": "False",
            "show_warning": "False",
            "theme": "minty",
            "invoice_recognition": "False",
            "enable_mkldnn": "False",
            "use_tensorrt": "False",
            "use_gpu": "False",
            "cpu_threshold": "65",
            "precision": "fp32",
            "font_family":"微软雅黑",
            "font_size": 10
        }
        self.load_all_setting()
        self.config["HISTORY"] = {}  # 空的历史记录
        self.save_config(False)
    def select_language(self):
        """通过Combobox选择语言"""
        lang_map = {
            "简体中文": "zh_CN",
            "English": "en_US"
        }
        try:
            self.selected_language = lang_map[self.language_combobox.get()]
        except:
            self.selected_language = self.config["DEFAULT"]["language"]

    def apply_language_change(self):
        """应用语言修改"""
        self.select_language()
        self.apply_font_settings()
        if self.selected_language != self.current_language:
            self.current_language = self.selected_language
            self.update_language()
    def create_widgets(self):
        # Create tab container
        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        self.load_tab_contents()  # 直接调用，移除多线程

    def pan_image(self, event):
        """处理图像平移"""
        if not hasattr(self, 'pan_start_x'):
            return

        # 计算移动增量
        delta_x = event.x - self.pan_start_x
        delta_y = event.y - self.pan_start_y

        # 更新总偏移量
        self.image_offset_x += delta_x
        self.image_offset_y += delta_y

        # 移动画布上的图像
        self.canvas.move("all", delta_x, delta_y)

        # 更新起始点为当前位置
        self.pan_start_x = event.x
        self.pan_start_y = event.y
    def load_tab_contents(self):
        # Create tabs
        edit_tab = ttk.Frame(self.notebook)
        recognize_tab = ttk.Frame(self.notebook)
        convert_tab = ttk.Frame(self.notebook)
        about_tab = ttk.Frame(self.notebook)
        settings_tab = ttk.Frame(self.notebook)
        optimize_tab = ttk.Frame(self.notebook)

        self.notebook.add(edit_tab, text=languages[self.current_language]["tab_edit"])
        self.notebook.add(recognize_tab, text=languages[self.current_language]["tab_recognize"])
        self.notebook.add(convert_tab, text=languages[self.current_language]["tab_convert"])
        self.notebook.add(optimize_tab, text=languages[self.current_language]["tab_optimize"])
        self.notebook.add(settings_tab, text=languages[self.current_language]["tab_settings"])
        self.notebook.add(about_tab, text=languages[self.current_language]["tab_about"])

        self.create_edit_tab(edit_tab)
        self.create_optimize_tab(optimize_tab)
        self.create_recognize_tab(recognize_tab)
        self.create_conversion_tab(convert_tab)
        self.create_settings_tab(settings_tab)
        self.create_about_tab(about_tab)
    def create_optimize_tab(self, tab):
        def on_enter(e):
            e.widget.config(bootstyle=SUCCESS)

        def on_leave(e):
            e.widget.config(bootstyle=(SUCCESS,OUTLINE))
        # Main container
        main_frame = tk.Frame(tab)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        # Left control panel
        control_frame = tk.Frame(main_frame)
        control_frame.pack(side=tk.LEFT, fill=tk.Y, padx=10)

        # File selection area
        file_frame = tk.Frame(control_frame)
        file_frame.pack(fill=tk.X, pady=5)
        self.optimize_file_label = ttk.Label(file_frame, text=languages[self.current_language]["select_image"])
        self.optimize_file_label.pack(side=tk.LEFT)
        self.optimize_file_path = ttk.Entry(file_frame, width=30)
        self.optimize_file_path.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        self.optimize_file_path.config(state='disabled')
        self.optimize_browse_button = ttk.Button(file_frame, text=languages[self.current_language]["browse"],
                                                command=self.browse_optimize_file, width=8)
        self.optimize_browse_button.pack(side=tk.LEFT)

        # Optimization options area
        self.options_frame = ttk.LabelFrame(control_frame, text=languages[self.current_language]["optimization_options"])
        self.options_frame.pack(fill=tk.X, pady=10)

        # Create optimization parameter widgets
        self.optimize_params = {}
        self.create_slider(self.options_frame, languages[self.current_language]["denoise_strength"], 0.0, 20.0, 0.0)
        self.create_slider(self.options_frame, languages[self.current_language]["sharpen_strength"], 0.0, 5.0, 0.0)
        self.create_slider(self.options_frame, languages[self.current_language]["brightness"], -100.0, 100.0, 0.0)
        self.create_slider(self.options_frame, languages[self.current_language]["contrast"], -100.0, 100.0, 0.0)
        self.create_slider(self.options_frame, languages[self.current_language]["edge_enhancement"], 0.0, 5.0, 0.0)

        # Operation buttons
        btn_frame = tk.Frame(control_frame)
        btn_frame.pack(fill=tk.X, pady=10)
        self.save_button = ttk.Button(btn_frame, text=languages[self.current_language]["save_result"], command=self.save_optimized_image,bootstyle=(SUCCESS,OUTLINE))
        self.save_button.pack(side=tk.LEFT, padx=5)
        self.reset_button = ttk.Button(btn_frame, text=languages[self.current_language]["reset"], command=self.reset_optimization,bootstyle=(SUCCESS,OUTLINE))
        self.reset_button.pack(side=tk.LEFT, padx=5)

        # Original preview
        self.before_preview_frame = ttk.LabelFrame(main_frame, text=languages[self.current_language]["original_preview"])
        self.before_preview_frame.pack(side=tk.TOP, anchor='ne', fill=tk.BOTH, expand=True)
        self.orig_preview_label = tk.Label(self.before_preview_frame)
        self.orig_preview_label.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # Processed result preview
        self.after_preview_frame = ttk.LabelFrame(main_frame, text=languages[self.current_language]["processed_preview"])
        self.after_preview_frame.pack(side=tk.TOP, anchor='ne', fill=tk.BOTH, expand=True)
        self.result_preview_label = tk.Label(self.after_preview_frame)
        self.result_preview_label.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)
        self.save_button.bind("<Enter>", on_enter)
        self.save_button.bind("<Leave>", on_leave)
        self.reset_button.bind("<Enter>", on_enter)
        self.reset_button.bind("<Leave>", on_leave)
    def create_edit_tab(self, tab):
        def on_enter(e):
            e.widget.config(bootstyle=SUCCESS)

        def on_leave(e):
            e.widget.config(bootstyle=(SUCCESS, OUTLINE))#SECONDARY
        # 主容器
        main_frame = tk.Frame(tab)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        # 左侧控制面板（优化布局）
        control_frame = tk.Frame(main_frame)
        control_frame.pack(side=tk.LEFT, fill=tk.Y, padx=10)

        # 文件选择区域（带原图尺寸显示）
        file_frame = tk.Frame(control_frame)
        file_frame.pack(fill=tk.X, pady=5)
        self.edit_file_label = ttk.Label(file_frame, text=languages[self.current_language]["edit_select_image"])
        self.edit_file_label.pack(side=tk.LEFT)
        self.edit_file_path = ttk.Entry(file_frame, width=25)
        self.edit_file_path.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        self.edit_file_path.config(state='disabled')
        self.browse_edit_btn = ttk.Button(file_frame, text=languages[self.current_language]["browse"],
                                         command=self.browse_edit_file, width=8)
        self.browse_edit_btn.pack(side=tk.LEFT)

        # 原图尺寸显示标签
        self.size_label = ttk.Label(control_frame, text="")
        self.size_label.pack(pady=5)
        self.mode_label = ttk.Label(control_frame,
                                    text=f"{languages[self.current_language]['current_mode']}: "
                                         f"{languages[self.current_language]['pan_mode' if self.current_mode == 'pan' else 'crop_mode']}",
                                    foreground="blue")
        self.mode_label.pack(pady=5)
        # 尺寸调整设置（带输入验证）
        self.resize_frame = ttk.LabelFrame(control_frame, text=languages[self.current_language]["resize_settings"])
        self.resize_frame.pack(fill=tk.X, pady=10)

        # 宽度输入（带数字验证）
        width_frame = tk.Frame(self.resize_frame)
        width_frame.pack(fill=tk.X, pady=2)
        self.width_label = ttk.Label(width_frame, text=languages[self.current_language]["width"], width=8)
        self.width_label.pack(side=tk.LEFT)
        self.width_var = tk.StringVar()
        vcmd = (control_frame.register(self.validate_number), '%P')
        ttk.Entry(width_frame, textvariable=self.width_var, width=10, validate="key", validatecommand=vcmd).pack(
            side=tk.LEFT)

        # 高度输入（带数字验证）
        height_frame = tk.Frame(self.resize_frame)
        height_frame.pack(fill=tk.X, pady=2)
        self.height_label = ttk.Label(height_frame, text=languages[self.current_language]["height"], width=8)
        self.height_label.pack(side=tk.LEFT)
        self.height_var = tk.StringVar()
        ttk.Entry(height_frame, textvariable=self.height_var, width=10, validate="key", validatecommand=vcmd).pack(
            side=tk.LEFT)

        # 保持比例复选框
        self.keep_ratio = tk.BooleanVar(value=True)
        self.ratio_check = ttk.Checkbutton(self.resize_frame,
                                           text=languages[self.current_language]["keep_ratio"],
                                           variable=self.keep_ratio)
        self.ratio_check.pack(anchor=tk.W, pady=5)

        # 右侧预览区域（优化性能）
        self.preview_frame = ttk.LabelFrame(main_frame, text=languages[self.current_language]["preview_title"])
        self.preview_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=10)
        self.edit_preview_label = tk.Label(self.preview_frame)
        self.edit_preview_label.pack(fill=tk.BOTH, expand=False)

        # 绑定带防抖的事件
        self.width_var.trace_add("write", lambda *_, f="width": self.debounced_preview(f))
        self.height_var.trace_add("write", lambda *_, f="height": self.debounced_preview(f))
        self._preview_job = None
        # 在尺寸调整设置区域后添加比例控件
        self.create_ratio_controls(self.resize_frame)
        # 裁剪控件
        self.crop_frame = ttk.LabelFrame(control_frame, text=languages[self.current_language]["crop_settings"])
        self.crop_frame.pack(fill=tk.X, pady=5)

        self.crop_coords = tk.StringVar()
        self.crop_label = ttk.Label(self.crop_frame, text=languages[self.current_language]["crop_coords"])
        self.crop_label.pack(side=tk.LEFT, fill=tk.X, expand = True)
        ttk.Entry(self.crop_frame, textvariable=self.crop_coords, width=20, state="readonly").pack(side=tk.LEFT)
        self.crop_btn = ttk.Button(self.crop_frame,
                                text=languages[self.current_language]["start_crop"],
                                command=self.start_crop_selection,
                                bootstyle=(SUCCESS, OUTLINE)
                                )
        self.crop_btn.pack(side=tk.RIGHT)
        # 添加Canvas用于裁剪选择
        self.canvas = tk.Canvas(self.preview_frame)
        self.canvas.pack(fill=tk.BOTH, expand=True)
        self.canvas.bind("<ButtonPress-1>", self.on_mouse_down)
        self.canvas.bind("<B1-Motion>", self.on_mouse_drag)
        self.canvas.bind("<ButtonRelease-1>", self.on_mouse_up)

        # 添加裁剪相关变量
        self.start_x = None
        self.start_y = None
        self.rect = None
        self.crop_enabled = False
        self.crop_area = None
        # 保存按钮（带加载状态）
        self.image_save_btn = ttk.Button(
            control_frame,
            text=languages[self.current_language]["save_edit"],
            command=self.save_edited_image,
            width=15,
            bootstyle=(SUCCESS, OUTLINE)
        )
        self.image_save_btn.pack(pady=10)
        self.image_save_btn.bind("<Enter>", on_enter)
        self.image_save_btn.bind("<Leave>", on_leave)
        self.crop_btn.bind("<Enter>", on_enter)
        self.crop_btn.bind("<Leave>", on_leave)
        # 绑定事件
        # 初始化平移相关变量
        self.pan_start_x = 0
        self.pan_start_y = 0
        self.image_offset_x = 0
        self.image_offset_y = 0
        self.setup_canvas_events()
        # 添加缩放提示标签
        self.zoom_tips_label = ttk.Label(control_frame,
                                         text=languages[self.current_language]["zoom_tips"],
                                         foreground="gray")
        self.zoom_tips_label.pack(pady=5)

        # 绑定双击事件重置缩放
        self.canvas.bind("<Double-Button-1>", self.reset_zoom)

    def zoom_image(self, event):
        """处理图片缩放（最终修正版）"""
        try:
            # 判断滚动方向
            scale_factor = 1.1 if (event.delta > 0 or event.num == 4) else 0.9

            # 限制缩放范围
            new_scale = self.zoom_scale * scale_factor
            if not (self.min_zoom <= new_scale <= self.max_zoom):
                return

            # 获取鼠标在画布上的绝对坐标
            canvas_x = self.canvas.canvasx(event.x)
            canvas_y = self.canvas.canvasy(event.y)

            # 获取当前图片坐标和尺寸
            img_x, img_y = self.canvas.coords(self.canvas_image_id)
            img_w = self.original_edit_image.width * self.zoom_scale
            img_h = self.original_edit_image.height * self.zoom_scale

            # 计算鼠标在图片上的相对位置（基于原始尺寸）
            rel_x = (canvas_x - img_x) / img_w
            rel_y = (canvas_y - img_y) / img_h

            # 应用新的缩放比例
            self.zoom_scale = new_scale

            # 计算新尺寸
            new_w = self.original_edit_image.width * self.zoom_scale
            new_h = self.original_edit_image.height * self.zoom_scale

            # 重置图片缩放（关键修改）
            self.canvas.delete(self.canvas_image_id)
            resized_img = self.original_edit_image.resize(
                (int(new_w), int(new_h)),
                Image.Resampling.LANCZOS
            )
            self.tk_image = ImageTk.PhotoImage(resized_img)
            self.canvas_image_id = self.canvas.create_image(img_x, img_y, anchor=tk.NW, image=self.tk_image)

            # 计算新的定位坐标
            new_img_x = canvas_x - (rel_x * new_w)
            new_img_y = canvas_y - (rel_y * new_h)

            # 移动图片到正确位置
            self.canvas.moveto(self.canvas_image_id, new_img_x, new_img_y)

            # 强制更新界面并检查边界
            self.canvas.update_idletasks()
            self._check_image_boundaries()

        except Exception as e:
            print(f"缩放错误: {str(e)}")

    def realtime_update_ui(self, result, current, total):
        """实时更新UI显示（带保留功能）"""
        # 清空进度显示（保留原有逻辑）
        self.update_progress(current, total)

        # 构建新内容
        content = f"\n=== 内容 {result.get('页码', 'N/A')} ===\n"
        content += f"识别状态: {result.get('状态', '未知')}\n"
        content += f"识别文本:\n{result.get('文本', '')}\n"

        # 更新文本框
        self.file_characters.config(state='normal')
        self.file_characters.insert(tk.END, content)
        self.file_characters.see(tk.END)
        self.file_characters.config(state='disabled')

        # 更新表格（原有逻辑不变）
        if self.invoice_recognition_var.get():
            validator = InvoiceValidator()
            validation = validator.validate_page(result["文本"])
            result.update(validation)
            self.update_results_table([result])
    # 添加moveto兼容方法
    def init_canvas_moveto(self):
        """为旧版Tkinter添加moveto兼容支持"""
        if not hasattr(self.canvas, 'moveto'):
            def canvas_moveto(canvas_self, item, x, y):
                current_coords = canvas_self.coords(item)
                if current_coords:
                    dx = x - current_coords[0]
                    dy = y - current_coords[1]
                    canvas_self.move(item, dx, dy)

            self.canvas.moveto = types.MethodType(canvas_moveto, self.canvas)
    def _check_image_boundaries(self):
        """边界检查（最终修正版）"""
        try:
            canvas_width = self.canvas.winfo_width()
            canvas_height = self.canvas.winfo_height()

            img_x, img_y = self.canvas.coords(self.canvas_image_id)
            img_w = self.original_edit_image.width * self.zoom_scale
            img_h = self.original_edit_image.height * self.zoom_scale

            # 动态计算边界阈值（增加10px缓冲）
            buffer = 10
            min_x = min(-buffer, canvas_width - img_w + buffer)
            max_x = max(buffer, canvas_width - buffer)
            min_y = min(-buffer, canvas_height - img_h + buffer)
            max_y = max(buffer, canvas_height - buffer)

            # 渐进式调整
            new_x = img_x
            new_y = img_y

            if img_x > max_x:
                new_x = max_x
            elif img_x < min_x:
                new_x = min_x

            if img_y > max_y:
                new_y = max_y
            elif img_y < min_y:
                new_y = min_y

            if (new_x, new_y) != (img_x, img_y):
                self.canvas.coords(self.canvas_image_id, new_x, new_y)

        except Exception as e:
            print(f"边界检查错误: {str(e)}")

    def reset_zoom(self, event=None):
        """重置缩放和位置"""
        if hasattr(self, 'original_edit_image'):
            self.zoom_scale = 1.0
            # 重置缩放并居中图片
            canvas_width = self.canvas.winfo_width()
            canvas_height = self.canvas.winfo_height()
            img_width = self.original_edit_image.width
            img_height = self.original_edit_image.height

            x = (canvas_width - img_width) / 2
            y = (canvas_height - img_height) / 2

            self.canvas.moveto(self.canvas_image_id, x, y)
            self._check_image_boundaries()
    def on_pan_start(self, event):
        """记录平移起始坐标"""
        self.pan_start_x = event.x
        self.pan_start_y = event.y

    def on_pan_move(self, event):
        """实时更新平移位置"""
        # 计算移动增量
        delta_x = event.x - self.pan_start_x
        delta_y = event.y - self.pan_start_y

        # 更新总偏移量
        self.image_offset_x += delta_x
        self.image_offset_y += delta_y

        # 移动画布上的图像
        self.canvas.move(self.canvas_image_id, delta_x, delta_y)

        # 更新起始点为当前位置
        self.pan_start_x = event.x
        self.pan_start_y = event.y

    def on_pan_end(self, event):
        """平移结束处理"""
        # 可在此添加边界检查逻辑
        self._check_image_boundaries()

    def create_loading_animation(self):
        """创建加载动画组件（修正版）"""
        self.loading_frame = ttk.Frame(self.root, relief='flat')
        self.loading_frame.place(relx=0.5, rely=0.5, anchor=tk.CENTER, width=200, height=100)

        # 加载文字
        self.loading_label = ttk.Label(
            self.loading_frame,
            text=languages[self.current_language]["loading"],
            font=('微软雅黑', 12)
        )
        self.loading_label.pack(pady=5)

        # 进度条
        self.loading_progress = ttk.Progressbar(
            self.loading_frame,
            mode='indeterminate',
            length=180,
            bootstyle=(STRIPED, SUCCESS)
        )
        self.loading_progress.pack(pady=5)

        # 初始隐藏
        self.loading_frame.lower()
        self.loading_frame.place_forget()

        # 添加旋转弧
        self.loading_canvas = tk.Canvas(
            self.loading_frame,
            width=40,
            height=40,
            highlightthickness=0
        )
        self.loading_canvas.pack()
        self.loading_arc = self.loading_canvas.create_arc(
            5, 5, 35, 35,
            start=0,
            extent=270,
            style=tk.ARC,
            outline="#4CAF50",
            width=3
        )
    def create_progress_animation(self):
        """创建渐变进度条"""
        self.style.configure("Custom.Horizontal.TProgressbar",
                             troughcolor='#e0e0e0',
                             background='linear-gradient(90deg, #4CAF50, #8BC34A)',
                             thickness=20)

        self.progress = ttk.Progressbar(self.root,
                                        orient=tk.HORIZONTAL,
                                        length=200,
                                        mode='determinate',
                                        style="Custom.Horizontal.TProgressbar")
        self.progress.pack(pady=10)

    def start_loading_animation(self):
        self.loading_frame.lift()
        self.loading_frame.place(relx=0.5, rely=0.5, anchor=tk.CENTER)
        self.loading_progress.start()
        self.animate_rotation(0)

    def stop_loading_animation(self):
        self.loading_progress.stop()
        self.loading_frame.place_forget()
        self.loading_canvas.delete("all")

    def animate_rotation(self, angle):
        if self.loading_frame.winfo_ismapped():
            self.loading_canvas.delete(self.loading_arc)
            self.loading_arc = self.loading_canvas.create_arc(
                5, 5, 35, 35,
                start=angle,
                extent=270,
                style=tk.ARC,
                outline="#4CAF50",
                width=3
            )
            self.root.after(50, lambda: self.animate_rotation((angle + 30) % 360))
    def debounced_preview(self, changed_field=None):
        """防抖优化版本（线程安全）"""
        # 确保 _preview_job 已初始化
        if not hasattr(self, '_preview_job'):
            self._preview_job = None

        # 取消之前的任务（线程安全）
        if self._preview_job is not None:
            try:
                self.root.after_cancel(self._preview_job)
            except (ValueError, TypeError):  # 捕获所有可能的无效 ID 异常
                pass
            finally:
                self._preview_job = None  # 强制重置状态

        # 设置新任务
        self._preview_job = self.root.after(300, lambda: self._execute_preview_update(changed_field))

    def _execute_preview_update(self, changed_field):
        """实际执行预览更新（确保主线程操作）"""
        try:
            # 获取当前输入值
            width_str = self.width_var.get().strip()
            height_str = self.height_var.get().strip()

            # 空值处理
            if not width_str and not height_str:
                return

            # 异步处理（防止主线程阻塞）
            threading.Thread(
                target=self._async_update_preview,
                args=(width_str, height_str),
                daemon=True
            ).start()
        except Exception as e:
            logging.error(f"Preview error: {str(e)}")

    def _async_update_preview(self, width_str, height_str):
        """后台处理图像"""
        try:
            # 解析数值
            width = int(width_str) if width_str else None
            height = int(height_str) if height_str else None

            # 处理图像（示例逻辑）
            resized_img = self.original_edit_image.resize((width, height), Image.Resampling.LANCZOS)

            # 主线程更新UI
            self.root.after(0, self.show_edit_image, resized_img)
        except ValueError:
            pass  # 忽略无效输入
        except Exception as e:
            self.root.after(0, lambda: messagebox.showerror("Error", str(e)))
    # 辅助方法
    def validate_number(self, P):
        """输入验证只允许数字"""
        if P.strip() == "" or P.isdigit():
            return True
        return False

    def browse_edit_file(self):
        filename = filedialog.askopenfilename(filetypes=[
            (languages[self.current_language]["image"], "*.jpg *.png *.bmp *.tif *.tiff"),
            (languages[self.current_language]["all_files"], "*.*")
        ])
        if filename:
            self.edit_file_path.config(state='normal')
            self.edit_file_path.delete(0, tk.END)
            self.edit_file_path.insert(0, filename)
            self.edit_file_path.config(state='disabled')
            self.load_edit_preview(filename)
            self.sync_image_path(filename)

    def load_edit_preview(self, path):
        try:
            # 异步加载防止界面卡顿
            def load_image():
                self.original_edit_image = Image.open(path)  # 确保先设置属性
                self.original_ratio = (self.original_edit_image.width,  # 在设置后访问
                                       self.original_edit_image.height)
                self.current_ratio = self.original_ratio
                self.root.after(0, self._update_size_display)
                self.root.after(0, self.show_edit_image, self.original_edit_image.copy())
                self.root.after(0, lambda: self.ratio_combobox.current(0))  # 在主线程操作UI

            threading.Thread(target=load_image, daemon=True).start()
        except Exception as e:
            messagebox.showerror(languages[self.current_language]["error"],
                                 languages[self.current_language]["load_image_failed"].format(error=str(e)))

    def _update_size_display(self):
        """显示原图尺寸"""
        w, h = self.original_edit_image.size
        self.size_label.config(text=languages[self.current_language]["current_size"].format(w, h))
        self.width_var.set(str(w))
        self.height_var.set(str(h))

    def show_edit_image(self, image):
        """通过 Canvas 显示图像"""
        # 清空 Canvas 原有内容
        self.canvas.delete("all")
        self.image_position = (0, 0)  # 初始位置设为画布原点
        self.zoom_scale = 1.0  # 重置缩放比例
        # 调整图像尺寸
        max_size = (2048, 2048)
        img_copy = image.copy()
        img_copy.thumbnail(max_size, Image.Resampling.LANCZOS)

        # 将图像转换为 Tkinter PhotoImage 并绘制到 Canvas
        self.tk_image = ImageTk.PhotoImage(img_copy)
        self.canvas_image_id = self.canvas.create_image(  # 保存图像ID
            0, 0,
            anchor=tk.NW,
            image=self.tk_image
        )

        # 记录实际显示尺寸（用于坐标转换）
        self.canvas.img_display_width = img_copy.width
        self.canvas.img_display_height = img_copy.height
        print("\n*** 图片加载调试 ***")
        print(f"原始尺寸: {image.size}")
        print(f"显示尺寸: {img_copy.size}")
        print(f"初始缩放比例: {self.zoom_scale}")
        print(f"初始位置: {self.image_position}")
    def update_resize_preview(self, changed_field):
        """优化后的尺寸预览方法"""
        try:
            if not hasattr(self, 'original_edit_image'):
                return

            # 获取输入值
            width_str = self.width_var.get().strip()
            height_str = self.height_var.get().strip()

            # 空值处理
            if not width_str and not height_str:
                self.show_edit_image(self.original_edit_image)
                return

            # 转换数值
            original_width, original_height = self.original_edit_image.size
            width = int(width_str) if width_str else None
            height = int(height_str) if height_str else None

            # 应用比例规则
            if self.keep_ratio.get():
                self.apply_aspect_ratio(changed_field)

            # 异步处理缩放
            if width and height:
                # 在后台线程处理缩放
                def resize_task():
                    resized_img = self.original_edit_image.resize(
                        (width, height),
                        Image.Resampling.LANCZOS
                    )
                    self.root.after(0, self.show_edit_image, resized_img)

                threading.Thread(target=resize_task, daemon=True).start()

        except ValueError:
            messagebox.showerror(languages[self.current_language]["error"],
                                 languages[self.current_language]["invalid_dimension"])

    def save_edited_image(self):
        try:
            # 初始化为原始图片
            resized_img = self.original_edit_image.copy()

            # 如果有有效的裁剪区域
            if hasattr(self, 'crop_area') and self.crop_area != (0, 0, 0, 0):
                # 获取原始图片尺寸
                img_w, img_h = self.original_edit_image.size

                # 转换Canvas坐标到实际图片坐标
                scale_x = img_w / self.canvas.img_display_width
                scale_y = img_h / self.canvas.img_display_height

                # 计算实际裁剪坐标
                x1 = int(self.crop_area[0] * scale_x)
                y1 = int(self.crop_area[1] * scale_y)
                x2 = int(self.crop_area[2] * scale_x)
                y2 = int(self.crop_area[3] * scale_y)

                # 确保坐标有效性
                x1 = max(0, min(x1, img_w))
                y1 = max(0, min(y1, img_h))
                x2 = max(x1, min(x2, img_w))
                y2 = max(y1, min(y2, img_h))

                # 执行裁剪
                resized_img = resized_img.crop((x1, y1, x2, y2))

            # 获取调整后的尺寸（如果有）
            try:
                new_width = int(self.width_var.get()) if self.width_var.get() else resized_img.width
                new_height = int(self.height_var.get()) if self.height_var.get() else resized_img.height
                resized_img = resized_img.resize((new_width, new_height), Image.Resampling.LANCZOS)
            except ValueError:
                pass

            # 保存最终图片
            save_path = filedialog.asksaveasfilename(defaultextension=".png")
            if save_path:
                resized_img.save(save_path)
                messagebox.showinfo(languages[self.current_language]["success"],
                                    languages[self.current_language]["image_saved"])
        except Exception as e:
            messagebox.showerror(languages[self.current_language]["error"],
                                 languages[self.current_language]["save_failed"].format(error=str(e)))
        # 动态获取实际显示的图像尺寸（需在加载图片时记录）
        img_w, img_h = self.original_edit_image.size
        scale_x = img_w / self.canvas.img_display_width  # 需在显示图片时记录实际显示尺寸
        scale_y = img_h / self.canvas.img_display_height

        # 修正坐标转换逻辑
        x1 = int(self.crop_area[0] * scale_x)
        y1 = int(self.crop_area[1] * scale_y)
        x2 = int(self.crop_area[2] * scale_x)
        y2 = int(self.crop_area[3] * scale_y)
    def create_slider(self, parent, label, from_, to, default):
        frame = tk.Frame(parent)
        frame.pack(fill=tk.X, pady=2)

        # Create label
        label_widget = ttk.Label(frame, text=label, width=13)
        label_widget.pack(side=tk.LEFT)

        # Create entry field
        entry_var = tk.DoubleVar(value=round(default, 2))
        entry = ttk.Entry(frame, textvariable=entry_var, width=5)
        entry.pack(side=tk.LEFT, padx=5)

        # Create slider
        slider_var = tk.DoubleVar(value=round(default, 2))
        slider = ttk.Scale(frame, from_=from_, to=to, variable=slider_var, orient=tk.HORIZONTAL)
        slider.pack(fill=tk.X, expand=True)

        # Bind slider and entry field values
        def update_entry(*args):
            value = slider_var.get()
            entry_var.set(value)

        def update_slider(*args):
            try:
                value = entry_var.get()
                if from_ <= value <= to:
                    slider_var.set(value)
                else:
                    messagebox.showwarning(
                        languages[self.current_language]["warning"],
                        languages[self.current_language]["input_range_error"].format(min=from_, max=to)
                    )
                    entry_var.set(slider_var.get())
            except ValueError:
                messagebox.showwarning(
                    languages[self.current_language]["warning"],
                    languages[self.current_language]["input_value_error"]
                )
                entry_var.set(slider_var.get())

        def update_preview(*args):
            self.apply_optimization()

        slider_var.trace("w", update_entry)
        entry_var.trace("w", update_slider)
        slider_var.trace("w", update_preview)

        # Store references to variables and label
        self.optimize_params[get_key_from_value(languages[self.current_language],label)] = {
            "slider_var": slider_var,
            "entry_var": entry_var,
            "label": label_widget
        }

        # Store label reference as an attribute of self
        setattr(self, f"{label}_label", label_widget)
    def browse_optimize_file(self):
        filename = filedialog.askopenfilename(filetypes=[
            (languages[self.current_language]["image"], "*.jpg *.png *.bmp *.tif *.tiff"),
            (languages[self.current_language]["all_files"], "*.*")
        ])
        if filename:
            self.optimize_file_path.config(state='normal')
            self.optimize_file_path.delete(0, tk.END)
            self.optimize_file_path.insert(0, filename)
            self.optimize_file_path.config(state='disabled')
            self.load_optimize_preview(filename)

    def load_optimize_preview(self, path):
        try:
            self.original_image = Image.open(path)
            self.processed_image = self.original_image.copy()

            # Show original image
            self.show_image(self.original_image, self.orig_preview_label)
            # Show processed image
            self.show_image(self.processed_image, self.result_preview_label)
        except Exception as e:
            messagebox.showerror(languages[self.current_language]["error"],
                                 languages[self.current_language]["load_image_failed"].format(error=str(e)))

    def show_image(self, image, label_widget):
        max_size = (400, 400)
        image.thumbnail(max_size, Image.Resampling.LANCZOS)
        photo = ImageTk.PhotoImage(image)
        label_widget.config(image=photo)
        label_widget.image = photo

    def save_success(self, save_path):
        """保存成功回调"""
        messagebox.showinfo(languages[self.current_language]["success"],
                            languages[self.current_language]["image_saved"])
        self.save_btn.config(state=tk.NORMAL, text=languages[self.current_language]["save_edit"])

    def save_error(self, error_msg):
        """保存失败回调"""
        messagebox.showerror(languages[self.current_language]["error"],
                             languages[self.current_language]["save_failed"].format(error=error_msg))
        self.save_btn.config(state=tk.NORMAL, text=languages[self.current_language]["save_edit"])

    def reset_save_button(self):
        """重置保存按钮状态"""
        self.save_btn.config(state=tk.NORMAL, text=languages[self.current_language]["save_edit"])

    def apply_optimization(self):
        if not hasattr(self, 'original_image'):
            messagebox.showwarning(languages[self.current_language]["warning"],
                                   languages[self.current_language]["select_image_first"])
            return
        try:
            # Convert to OpenCV format
            img = cv2.cvtColor(np.array(self.original_image), cv2.COLOR_RGB2BGR)
            denoise_strength = self.optimize_params["denoise_strength"]["entry_var"].get()
            if denoise_strength > 0:
                ksize = int(denoise_strength // 5 * 2 + 1)
                if ksize % 2 == 0:  # Ensure odd kernel size
                    ksize += 1
                if ksize < 3:
                    ksize = 3
                img = cv2.medianBlur(img, ksize)

            # Sharpening
            sharpen_strength = self.optimize_params["sharpen_strength"]["entry_var"].get()
            if sharpen_strength > 0:
                # Use more complex sharpening algorithm
                kernel = np.array([[-1, -1, -1],
                                   [-1, 9 + sharpen_strength, -1],
                                   [-1, -1, -1]])
                img = cv2.filter2D(img, -1, kernel)

            # Brightness and contrast
            contrast = self.optimize_params["contrast"]["entry_var"].get()
            brightness = self.optimize_params["brightness"]["entry_var"].get()
            alpha = 1 + contrast / 100
            beta = brightness
            img = cv2.convertScaleAbs(img, alpha=alpha, beta=beta)

            # Edge enhancement
            edge_strength = self.optimize_params["edge_enhancement"]["entry_var"].get()
            if edge_strength > 0:
                img = cv2.detailEnhance(img, sigma_s=10, sigma_r=edge_strength / 10)

            # Convert back to PIL format
            self.processed_image = Image.fromarray(cv2.cvtColor(img, cv2.COLOR_BGR2RGB))
            self.show_image(self.processed_image, self.result_preview_label)

        except Exception as e:
            messagebox.showerror(languages[self.current_language]["error"],
                languages[self.current_language]["optimization_failed"].format(error=str(e)))

    def save_optimized_image(self):
        if not hasattr(self, 'processed_image'):
            return
        filename = filedialog.asksaveasfilename(
            defaultextension=".png",
            filetypes=[("PNG", "*.png"), ("JPG", "*.jpg"),
                       (languages[self.current_language]["all_files"], "*.*")]
        )
        if filename:
            try:
                self.processed_image.save(filename)
                messagebox.showinfo(languages[self.current_language]["success"],
                                    languages[self.current_language]["image_saved"])
            except Exception as e:
                messagebox.showerror(languages[self.current_language]["error"],
                                     languages[self.current_language]["save_failed"].format(error=str(e)))

    def reset_optimization(self):
        if hasattr(self, 'original_image'):
            self.processed_image = self.original_image.copy()
            self.show_image(self.processed_image, self.result_preview_label)
        for param in self.optimize_params.values():
            if isinstance(param, tk.DoubleVar):
                param.set(0)

    def create_recognize_tab(self, tab):
        def on_enter(e):
            e.widget.config(bootstyle=SUCCESS)

        def on_leave(e):
            e.widget.config(bootstyle=(SUCCESS, OUTLINE))
        # Main container
        main_frame = tk.Frame(tab)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        # Left input area
        left_frame = tk.Frame(main_frame)
        left_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # Right result area
        right_frame = tk.Frame(main_frame)
        right_frame.pack(side=tk.RIGHT, fill=tk.BOTH, padx=10)

        # File path section
        file_frame = tk.Frame(left_frame)
        file_frame.pack(fill=tk.X, pady=5)
        self.file_path_label = ttk.Label(file_frame, text=languages[self.current_language]["file_path"])
        self.file_path_label.pack(side=tk.LEFT)
        self.file_path = ttk.Entry(file_frame, width=50)
        self.file_path.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        self.file_path.config(state='disabled')
        self.browse_button = ttk.Button(file_frame, text=languages[self.current_language]["browse"], command=self.browse_file, width=8)
        self.browse_button.pack(side=tk.LEFT)

        # Settings section
        self.optimize_settings_frame = ttk.LabelFrame(left_frame, text=languages[self.current_language]["settings"])
        self.optimize_settings_frame.pack(fill=tk.X, pady=5)

        # Excel name
        self.excel_name_label = ttk.Label(self.optimize_settings_frame, text=languages[self.current_language]["excel_name"])
        self.excel_name_label.pack(anchor=tk.W)
        self.excel_name = ttk.Entry(self.optimize_settings_frame)
        self.excel_name.pack(fill=tk.X, pady=2)

        # Optimization options
        self.optimize_mode_label = ttk.Label(self.optimize_settings_frame, text=languages[self.current_language]["recognition_optimize"])
        self.optimize_mode_label.pack(anchor=tk.W)
        self.optimize_mode = ttk.Combobox(self.optimize_settings_frame, values=self.optimize_list, state="readonly")
        self.optimize_mode.pack(fill=tk.X, pady=2)
        self.optimize_mode.set(languages[self.current_language]["optimize_none"])

        # OCR module
        self.ocr_module_label = ttk.Label(self.optimize_settings_frame, text=languages[self.current_language]["ocr_module"])
        self.ocr_module_label.pack(anchor=tk.W)
        self.ocr_module = ttk.Combobox(self.optimize_settings_frame, values=self.ocr_module_list, state="readonly")
        self.ocr_module.pack(fill=tk.X, pady=2)
        self.ocr_module.set(languages[self.current_language]["ocr_paddle"])

        # Operation buttons area
        btn_frame = tk.Frame(left_frame)
        btn_frame.pack(fill=tk.X, pady=10)
        self.start_button = ttk.Button(btn_frame,
            text=languages[self.current_language]["start_recognition"],
            command=self.main,
            bootstyle=(SUCCESS, OUTLINE)
            )
        self.start_button.pack(side=tk.LEFT, padx=5)
        self.export_button = ttk.Button(btn_frame, text=languages[self.current_language]["export_to_a1"], command=self.write_to_excel,bootstyle=(SUCCESS, OUTLINE))
        self.export_button.pack(side=tk.LEFT, padx=5)

        # Row and column settings
        rowcol_frame = tk.Frame(left_frame)
        rowcol_frame.pack(fill=tk.X)
        self.row_label = ttk.Label(rowcol_frame, text=languages[self.current_language]["row"])
        self.row_label.pack(side=tk.LEFT)
        self.row_entry = ttk.Entry(rowcol_frame, width=8)
        self.row_entry.pack(side=tk.LEFT, padx=5)
        self.col_label = ttk.Label(rowcol_frame, text=languages[self.current_language]["column"])
        self.col_label.pack(side=tk.LEFT)
        self.col_entry = ttk.Entry(rowcol_frame, width=8)
        self.col_entry.pack(side=tk.LEFT, padx=5)
        self.export_pos_button = ttk.Button(rowcol_frame, text=languages[self.current_language]["export_to_position"],
                                           command=self.write_to_excel_specified,bootstyle=(SUCCESS, OUTLINE))
        self.export_pos_button.pack(side=tk.LEFT)
        # 在操作按钮区域下方添加保留内容选项
        retain_frame = ttk.Frame(left_frame)
        retain_frame.pack(fill=tk.X, pady=5)
        ttk.Checkbutton(
            retain_frame,
            text=languages[self.current_language]["保留识别内容"],  # 需要添加到语言字典
            variable=self.retain_text_var
        ).pack(side=tk.LEFT)
        # Right result area
        self.result_frame = ttk.LabelFrame(right_frame, text=languages[self.current_language]["recognition_result"])
        self.result_frame.pack(fill=tk.BOTH, expand=True)

        self.file_characters = tk.Text(self.result_frame, wrap=tk.WORD)
        self.file_characters.pack(side=tk.TOP,fill=tk.BOTH, expand=True, padx=5, pady=5)
        self.file_characters.config(state='disabled')

        btn_bottom_frame = tk.Frame(self.result_frame)
        btn_bottom_frame.pack(fill=tk.X, pady=5)
        self.copy_button = ttk.Button(btn_bottom_frame,
            text=languages[self.current_language]["copy_content"],
            command = self.copy_to_clipboard,
            bootstyle=(SUCCESS, OUTLINE)
        )
        self.copy_button.pack(side=tk.LEFT)
        self.clear_button = ttk.Button(btn_bottom_frame,
            text=languages[self.current_language]["clear_content"],
            command = self.clear_text,
            bootstyle=(SUCCESS, OUTLINE)
            )
        self.clear_button.pack(side=tk.RIGHT)

        # Progress bar
        self.progress = ttk.Progressbar(self.root, mode='indeterminate')
        self.progress.pack(fill=tk.X, padx=10, pady=5)

        # Preview frame
        self.Key_preview_frame = ttk.LabelFrame(left_frame, text=languages[self.current_language]["file_preview"])
        self.Key_preview_frame.pack(fill=tk.BOTH, expand=True, pady=10)

        self.preview_label = tk.Label(self.Key_preview_frame)
        self.preview_label.pack(fill=tk.BOTH, expand=True)

        self.clear_history_frame = ttk.Frame(btn_bottom_frame)
        self.clear_history_frame.pack(side=tk.LEFT, padx=5)

        self.clear_selected_btn = ttk.Button(
            self.clear_history_frame,
            text=languages[self.current_language]["clear_selected"],
            command=self.clear_selected_history,
            bootstyle=(SUCCESS, OUTLINE)
        )
        self.clear_selected_btn.pack(side=tk.LEFT)
        # 添加结果表格
        if self.current_language == "zh_CN":
            columns = ["页码", "发票号码","开票日期","状态"]
        else:
            columns = ["Page","Invoice Number","Invoice Date","Status"]
        self.result_tree = ttk.Treeview(
            self.result_frame,
            columns=columns,
            show="headings",
            height=8
        )

        # 设置列宽
        col_widths = [60, 150, 140, 240]
        for col, width in zip(columns, col_widths):
            self.result_tree.heading(col, text=col)
            self.result_tree.column(col, width=width, anchor='center')

        # 添加滚动条
        scrollbar = ttk.Scrollbar(self.result_frame, orient="vertical",
                                  command=self.result_tree.yview)
        self.result_tree.configure(yscrollcommand=scrollbar.set)

        self.result_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=5, pady=5)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y, padx=(0, 5), pady=5)
        self.result_frame.pack_propagate(False)  # 禁止自动调整大小
        self.result_frame.config(width=650)  # 固定高度
        self.progress_frame = ttk.Frame(btn_bottom_frame)
        self.progress_frame.pack(side=tk.LEFT, padx=10)

        # 进度百分比标签
        self.progress_label = ttk.Label(self.progress_frame, text="0%")
        self.progress_label.pack(side=tk.TOP)

        # 进度条
        self.progress = ttk.Progressbar(self.progress_frame,
                                        mode='determinate',
                                        maximum=100,
                                        bootstyle=(SUCCESS, STRIPED))
        self.progress.pack(side=tk.TOP)

        # 数量状态标签
        self.count_label = ttk.Label(self.progress_frame, text="0/0")
        self.count_label.pack(side=tk.TOP)
        # Bind file path change event
        self.clear_selected_btn.bind("<Enter>", on_enter)
        self.clear_selected_btn.bind("<Leave>", on_leave)
        self.copy_button.bind("<Enter>", on_enter)
        self.copy_button.bind("<Leave>", on_leave)
        self.clear_button.bind("<Enter>", on_enter)
        self.clear_button.bind("<Leave>", on_leave)
        self.start_button.bind("<Enter>", on_enter)
        self.start_button.bind("<Leave>", on_leave)
        self.export_button.bind("<Enter>", on_enter)
        self.export_button.bind("<Leave>", on_leave)
        self.export_pos_button.bind("<Enter>", on_enter)
        self.export_pos_button.bind("<Leave>", on_leave)
    # 在ClassMain类中添加进度更新方法：
    def update_progress(self, current, total):
        progress = int((current / total) * 100)
        self.progress["value"] = progress
        self.progress_label.config(text=f"{progress}%")
        self.count_label.config(text=f"{current}/{total}")
        self.root.update_idletasks()
    def update_results_table(self, results):
        """更新结果表格"""

        #self.result_tree.delete(*self.result_tree.get_children())

        for result in results:
            values = []
            number = result.get("发票号码", "无")
            status = "有效发票"

            # 检查重复
            is_duplicate, timestamp = InvoiceValidator().check_duplicate(number, self.history,self.warning_prompt_var.get())
            if is_duplicate:
                status = "重复发票（首次验证时间：{}）".format(timestamp)

            # 记录新发票
            if not is_duplicate and number != "无":
                self.history[number] = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                self.save_config(False)  # 静默保存

            # 构建表格数据
            for col in ["页码", "发票号码","开票日期", "状态"]:
                if col == "状态":
                    values.append(status)
                else:
                    values.append(result.get(col, "无"))

            tag = 'duplicate' if is_duplicate else 'valid' if "有效" in status else 'invalid'
            self.result_tree.insert("", "end", values=values, tags=(tag,))

        # 设置行颜色
        self.result_tree.tag_configure('valid', background='#e6ffe6')
        self.result_tree.tag_configure('invalid', background='#ffe6e6')
        self.result_tree.tag_configure('duplicate', background='#fff3cd')

    # 在ClassMain类中添加新方法
    def save_invoice_validation(self):
        if not self.extracted_info:
            messagebox.showwarning(languages[self.current_language]["warning"],
                                   languages[self.current_language]["no_invoice_info"])
            return

        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx"), (languages[self.current_language]["all_files"], "*.*")]
        )
        if not file_path:
            return

        try:
            wb = Workbook()
            ws = wb.active
            ws.append([languages[self.current_language]["field"],
                       languages[self.current_language]["value"]])

            # 添加提取的信息
            for key, value in self.extracted_info.items():
                ws.append([key, value])

            # 添加验证结果
            ws.append([languages[self.current_language]["validation_result"],
                       self.validation_result])

            wb.save(file_path)
            messagebox.showinfo(languages[self.current_language]["success"],
                                languages[self.current_language]["invoice_saved"])
        except Exception as e:
            messagebox.showerror(languages[self.current_language]["error"],
                                 f"{languages[self.current_language]['save_failed']}: {str(e)}")
    def update_preview(self, event=None):
        """Update file preview"""
        file_path = self.file_path.get()
        if not file_path:
            return

        try:
            # Clear preview area
            self.preview_label.config(image='')

            # Handle PDF preview
            if file_path.lower().endswith('.pdf'):
                doc = fitz.open(file_path)
                page = doc.load_page(0)
                pix = page.get_pixmap()
                img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
                doc.close()
            else:
                # Handle image preview
                img = Image.open(file_path)

            # Adjust image size to fit preview area
            max_size = (400, 500)
            img.thumbnail(max_size, Image.Resampling.LANCZOS)
            photo = ImageTk.PhotoImage(img)

            # Update preview
            self.preview_label.config(image=photo)
            self.preview_label.image = photo  # Keep reference

        except Exception as e:
            logging.error(f"Preview failed: {str(e)}")
            tk.messagebox.showwarning(languages[self.current_language]["warning"],
                                      languages[self.current_language]["load_image_failed"].format(error=str(e)))

    def create_conversion_tab(self, tab):
        def on_enter(e):
            e.widget.config(bootstyle=SUCCESS)

        def on_leave(e):
            e.widget.config(bootstyle=(SUCCESS, OUTLINE))
        # File path area
        convert_frame = tk.Frame(tab)
        convert_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        # Conversion options section
        conversion_frame = tk.Frame(convert_frame)
        conversion_frame.pack(fill=tk.X, pady=5)
        self.conversion_label = ttk.Label(conversion_frame, text=languages[self.current_language]["select_conversion"])
        self.conversion_label.pack(side=tk.LEFT)

        conversion_values = [
            languages[self.current_language]["conversion_all"],
            languages[self.current_language]["conversion_pdf_to_image"],
            languages[self.current_language]["conversion_image_to_pdf"],
            languages[self.current_language]["conversion_txt_to_pdf"],
            languages[self.current_language]["conversion_merge_pdf"],
            languages[self.current_language]["conversion_info_only"]
        ]

        self.conversion_options = ttk.Combobox(conversion_frame, values=conversion_values, state="readonly")
        self.conversion_options.pack(side=tk.LEFT, fill=tk.X, pady=2)
        self.conversion_options.bind("<<ComboboxSelected>>", self.update_conversion_widgets)
        self.conversion_options.current(0)  # Default to first option

        # PDF to PNG/JPEG/TIFF area
        self.pdf_to_image_frame = ttk.LabelFrame(convert_frame, text=languages[self.current_language]["pdf_to_image"])
        self.pdf_to_image_frame.pack(fill=tk.X, pady=5)
        self.pdf_to_image_path = ttk.Entry(self.pdf_to_image_frame, width=50)
        self.pdf_to_image_path.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        self.pdf_to_image_path.config(state='disabled')
        ttk.Button(self.pdf_to_image_frame, text=languages[self.current_language]["browse"],
                  command=self.browse_pdf_to_image, width=8).pack(side=tk.LEFT)
        self.output_image_format = ttk.Combobox(self.pdf_to_image_frame,
                                                values=["PNG", "JPEG", "TIFF"],
                                                state="readonly")
        self.output_image_format.pack(side=tk.LEFT, padx=5)
        self.output_image_format.current(0)
        ttk.Button(self.pdf_to_image_frame, text=languages[self.current_language]["start_conversion"],
                  command=self.convert_pdf_to_image, width=8).pack(side=tk.LEFT)

        # Image to PDF area
        self.image_to_pdf_frame = ttk.LabelFrame(convert_frame, text=languages[self.current_language]["image_to_pdf"])
        self.image_to_pdf_frame.pack(fill=tk.X, pady=5)
        self.image_to_pdf_path = ttk.Entry(self.image_to_pdf_frame, width=50)
        self.image_to_pdf_path.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        self.image_to_pdf_path.config(state='disabled')
        ttk.Button(self.image_to_pdf_frame, text=languages[self.current_language]["browse"],
                  command=self.browse_image_to_pdf, width=8).pack(side=tk.LEFT)
        self.image_to_pdf_multiple = tk.IntVar()
        tk.Checkbutton(self.image_to_pdf_frame, text=languages[self.current_language]["merge_multiple_images"],
                       variable=self.image_to_pdf_multiple).pack(side=tk.LEFT, padx=5)
        ttk.Button(self.image_to_pdf_frame, text=languages[self.current_language]["start_conversion"],
                  command=self.convert_image_to_pdf, width=8).pack(side=tk.LEFT)

        # TXT to PDF area
        self.txt_to_pdf_frame = ttk.LabelFrame(convert_frame, text=languages[self.current_language]["txt_to_pdf"])
        self.txt_to_pdf_frame.pack(fill=tk.X, pady=5)
        self.txt_to_pdf_path = ttk.Entry(self.txt_to_pdf_frame, width=50)
        self.txt_to_pdf_path.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        self.txt_to_pdf_path.config(state='disabled')
        ttk.Button(self.txt_to_pdf_frame, text=languages[self.current_language]["browse_txt"],
                  command=self.browse_txt_to_pdf, width=8).pack(side=tk.LEFT)
        ttk.Button(self.txt_to_pdf_frame, text=languages[self.current_language]["start_conversion"],
                  command=self.convert_txt_to_pdf, width=8).pack(side=tk.LEFT)

        # PDF merge area
        self.pdf_merge_frame = ttk.LabelFrame(convert_frame, text=languages[self.current_language]["merge_pdf"])
        self.pdf_merge_frame.pack(fill=tk.X, pady=5)
        self.pdf_files = []
        self.pdf_listbox = tk.Listbox(self.pdf_merge_frame, selectmode=tk.MULTIPLE)
        self.pdf_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=5, pady=5)
        button_frame = tk.Frame(self.pdf_merge_frame)
        button_frame.pack(side=tk.LEFT, fill=tk.Y)
        # Add buttons to button_frame, arranged vertically
        ttk.Button(button_frame, text=languages[self.current_language]["add_file"],
                  command=self.add_pdf_files, width=8).pack(side=tk.TOP, pady=2)
        ttk.Button(button_frame, text=languages[self.current_language]["delete_selected"],
                  command=self.remove_pdf_files, width=8).pack(side=tk.TOP, pady=2)
        ttk.Button(button_frame, text=languages[self.current_language]["start_merge"],
                  command=self.merge_pdfs, width=8).pack(side=tk.TOP, pady=2)

        # Display area
        self.display_frame = ttk.LabelFrame(convert_frame, text=languages[self.current_language]["conversion_result"])
        self.display_frame.pack(fill=tk.BOTH, expand=True, pady=5)
        self.display_text = scrolledtext.ScrolledText(self.display_frame, wrap=tk.WORD, state="disabled")
        self.display_text.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

    def create_about_tab(self, tab):
        def on_enter(e):
            e.widget.config(bootstyle=SUCCESS)

        def on_leave(e):
            e.widget.config(bootstyle=(SUCCESS, OUTLINE))
        # About tab area
        about_frame = tk.Frame(tab)
        about_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        self.developer_label = ttk.Label(about_frame, text=languages[self.current_language]["developer"])
        self.developer_label.pack(side=tk.TOP, pady=10)

        self.dev_date_label = ttk.Label(about_frame, text=languages[self.current_language]["dev_date"])
        self.dev_date_label.pack(side=tk.TOP, pady=10)

        self.version_label = ttk.Label(about_frame, text=languages[self.current_language]['current_version'].format(version=latest_version))
        self.version_label.pack(side=tk.TOP, padx=10)

        self.update_button = ttk.Button(about_frame, text=languages[self.current_language]["check_update"],
                                        command=self.check_for_updates, width=8)
        self.update_button.pack(side=tk.TOP, padx=10)

        # Add language selection

        self.log_frame = ttk.LabelFrame(about_frame, text=languages[self.current_language]["update_log"])
        self.log_frame.pack(side=tk.TOP, expand=True, fill=tk.BOTH, padx=10, pady=10)

        self.logtext = tk.Text(self.log_frame, height=10, width=55, background="white")
        self.logtext.pack(side=tk.TOP, expand=True, fill=tk.BOTH)
        self.logtext.config(state='normal')
        self.logtext.delete(1.0, tk.END)
        log = ch_log if self.current_language == 'zh_CN' else en_log
        for log_entry in log:
            self.logtext.insert(tk.END, log_entry + '\n\n')
        self.logtext.config(state='disabled')

        # Create scrollbar
        scrollbar = ttk.Scrollbar(self.logtext, command=self.logtext.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # Configure text box and scrollbar linkage
        self.logtext.configure(yscrollcommand=scrollbar.set)
        self.logtext.config(state='disable')

        self.copyright_label = ttk.Label(about_frame, text=languages[self.current_language]["copyright"])
        self.copyright_label.pack(side=tk.BOTTOM, pady=10)

    def update_conversion_widgets(self, event):
        selected_option = self.conversion_options.get()
        self.reset_conversion_widgets()

        # Map selected option to internal option
        option_map = {
            languages[self.current_language]["conversion_all"]: "all",
            languages[self.current_language]["conversion_pdf_to_image"]: "pdf_to_image",
            languages[self.current_language]["conversion_image_to_pdf"]: "image_to_pdf",
            languages[self.current_language]["conversion_txt_to_pdf"]: "txt_to_pdf",
            languages[self.current_language]["conversion_merge_pdf"]: "merge_pdf",
            languages[self.current_language]["conversion_info_only"]: "info_only"
        }

        internal_option = option_map.get(selected_option, "all")

        if internal_option == "all":
            self.pdf_to_image_frame.pack(fill=tk.X, pady=5)
            self.image_to_pdf_frame.pack(fill=tk.X, pady=5)
            self.txt_to_pdf_frame.pack(fill=tk.X, pady=5)
            self.pdf_merge_frame.pack(fill=tk.X, pady=5)
        elif internal_option == "pdf_to_image":
            self.pdf_to_image_frame.pack(fill=tk.X, pady=5)
        elif internal_option == "image_to_pdf":
            self.image_to_pdf_frame.pack(fill=tk.X, pady=5)
        elif internal_option == "txt_to_pdf":
            self.txt_to_pdf_frame.pack(fill=tk.X, pady=5)
        elif internal_option == "merge_pdf":
            self.pdf_merge_frame.pack(fill=tk.X, pady=5)
        else:
            self.pdf_to_image_frame.pack_forget()
            self.image_to_pdf_frame.pack_forget()
            self.txt_to_pdf_frame.pack_forget()
            self.pdf_merge_frame.pack_forget()

        self.display_frame.pack(fill=tk.BOTH, expand=True, pady=5)

    def reset_conversion_widgets(self):
        self.pdf_to_image_frame.pack_forget()
        self.image_to_pdf_frame.pack_forget()
        self.txt_to_pdf_frame.pack_forget()
        self.pdf_merge_frame.pack_forget()
        self.display_frame.pack_forget()

    def clear_text(self):
        self.file_characters.config(state='normal')
        self.file_characters.delete("1.0", tk.END)
        self.file_characters.config(state='disabled')

    def copy_to_clipboard(self):
        self.file_characters.config(state='normal')
        content = self.file_characters.get("1.0", "end-1c")
        self.file_characters.config(state='disabled')
        if content:
            self.root.clipboard_clear()
            self.root.clipboard_append(content)
            tk.messagebox.showinfo(languages[self.current_language]["success"],
                                   languages[self.current_language]["copied_to_clipboard"])
        else:
            tk.messagebox.showwarning(languages[self.current_language]["warning"],
                                      languages[self.current_language]["nothing_to_copy"])

    def update_info_display(self, info, validation_msg):
        """Update information display area"""
        self.result_tree.delete(*self.result_tree.get_children())
        number = info.get("发票号码", "")
        history_key = number if number else ""
        # 添加时间戳信息
        if history_key in self.history:
            info["验证时间"] = self.history[history_key]
        # Define field display order
        field_order = ["发票类型",  "发票号码","发票代码", "开票日期", "税率", "金额", "税额", "校验码"]

        # Display fields in order
        for key in field_order:
            if key in info:
                self.result_tree.insert("", "end", values=(key, info[key]))

        # Update validation result label, set color based on result
        if "发现重复记录" in validation_msg or "Invoice Duplicate Check" in validation_msg:
            self.validation_label.config(text=validation_msg, fg="orange")
        elif "发票有效" in validation_msg or "Invoice valid" in validation_msg:
            self.validation_label.config(text=validation_msg, fg="green")
        else:
            self.validation_label.config(text=validation_msg, fg="red")

    def write_to_excel_specified(self):
        output_name = self.excel_name.get()
        data = self.file_characters.get("1.0", "end-1c")
        if not output_name:
            tk.messagebox.showerror(languages[self.current_language]["error"],
                                    languages[self.current_language]["empty_excel_name"])
            return

        try:
            row = int(self.row_entry.get())
            col = int(self.col_entry.get())
            if row <= 0 or col <= 0:
                raise ValueError("Row and column must be greater than 0")
        except ValueError as e:
            tk.messagebox.showerror(languages[self.current_language]["error"],
                                    languages[self.current_language]["invalid_row_col"].format(error=str(e)))
            return

        workbook = Workbook()
        sheet = workbook.active
        sheet.cell(row=row, column=col, value=data)

        try:
            workbook.save(output_name + '.xlsx')
            tk.messagebox.showinfo(languages[self.current_language]["success"],
                                   languages[self.current_language]["excel_saved"])
        except Exception as e:
            tk.messagebox.showerror(languages[self.current_language]["error"],
                                    languages[self.current_language]["excel_save_failed"].format(error=str(e)))

    def browse_file(self):
        filename = filedialog.askopenfilename(filetypes=[
            (languages[self.current_language]["image_pdf"], "*.jpg *.png *.bmp *.tif *.tiff *.pdf"),
            (languages[self.current_language]["all_files"], "*.*")
        ])
        if filename:
            self.file_path.config(state='normal')
            self.file_path.delete(0, tk.END)
            self.file_path.insert(0, filename)
            self.file_path.config(state='disabled')
            self.update_preview()
    def browse_pdf_to_image(self):
        filename = filedialog.askopenfilename(filetypes=[
            ("PDF", "*.pdf"),
            (languages[self.current_language]["all_files"], "*.*")
        ])
        if filename:
            self.pdf_to_image_path.config(state='normal')
            self.pdf_to_image_path.delete(0, tk.END)
            self.pdf_to_image_path.insert(0, filename)
            self.pdf_to_image_path.config(state='disabled')

    def browse_image_to_pdf(self):
        filenames = filedialog.askopenfilenames(filetypes=[
            (languages[self.current_language]["image"], "*.jpg *.png *.bmp *.tif *.tiff"),
            (languages[self.current_language]["all_files"], "*.*")
        ])
        if filenames:
            self.image_to_pdf_path.config(state='normal')
            self.image_to_pdf_path.delete(0, tk.END)
            self.image_to_pdf_path.insert(0, ', '.join(filenames))
            self.image_to_pdf_path.config(state='disabled')

    def browse_txt_to_pdf(self):
        filename = filedialog.askopenfilename(filetypes=[
            ("TXT", "*.txt"),
            (languages[self.current_language]["all_files"], "*.*")
        ])
        if filename:
            self.txt_to_pdf_path.config(state='normal')
            self.txt_to_pdf_path.delete(0, tk.END)
            self.txt_to_pdf_path.insert(0, filename)
            self.txt_to_pdf_path.config(state='disabled')

    def add_pdf_files(self):
        files = filedialog.askopenfilenames(filetypes=[
            ("PDF", "*.pdf"),
            (languages[self.current_language]["all_files"], "*.*")
        ])
        for file in files:
            self.pdf_listbox.insert(tk.END, file)
            self.pdf_files.append(file)

    def remove_pdf_files(self):
        selected_indices = self.pdf_listbox.curselection()
        if selected_indices:
            for index in sorted(selected_indices, reverse=True):
                self.pdf_listbox.delete(index)
                del self.pdf_files[index]

    def long_running_task(self):
        """耗时任务"""
        try:
            self.start_loading_animation()  # 显示加载动画
            self.init_ocr_engine()
            image_path = self.file_path.get()
            if not image_path:
                tk.messagebox.showerror(languages[self.current_language]["error"],
                                        languages[self.current_language]["empty_file_path"])
                return

            # Disable buttons to prevent repeated clicks
            self.toggle_buttons_state(disabled=True)
            # Start progress bar
            self.progress.start()
            # Start background thread for OCR
            threading.Thread(target=self.run_ocr_in_background, args=(image_path,), daemon=True).start()
        except:
            pass
    def main(self):
        self.start_loading_animation()
        # 在后台线程执行耗时操作
        threading.Thread(target=self.long_running_task, daemon=True).start()
    def run_ocr_in_background(self, image_path):
        nowtimer = time.time()
        print("start ocr")
        try:
            # 新增：根据复选框状态决定是否清空内容
            if not self.retain_text_var.get():
                self.root.after(0, lambda: [
                    self.result_tree.delete(*self.result_tree.get_children()),
                    self.file_characters.config(state='normal'),
                    self.file_characters.delete(1.0, tk.END),
                    self.file_characters.config(state='disabled')
                ])
            #self.root.after(0, lambda: self.result_tree.delete(*self.result_tree.get_children()))
            # 规范化文件路径处理
            image_path = os.path.normcase(image_path)

            # 获取优化模式索引
            optimize_mode_text = self.optimize_mode.get()
            optimize_mode = self.optimize_list.index(optimize_mode_text)

            # 启动加载动画
            self.root.after(0, self.start_loading_animation)

            # 禁用操作按钮防止重复点击
            self.root.after(0, self.toggle_buttons_state, True)

            # 根据文件类型调用不同处理方法
            if image_path.lower().endswith('.pdf'):
                # 调用 extract_text_from_pdf 处理 PDF
                results = FileProcessor.extract_text_from_pdf(
                    pdf_path=image_path,
                    ocr_engine=self.ocr_engine,
                    image_processor=self.image_processor,
                    optimize_mode=optimize_mode,
                    cpu_threshold=self.config.getint("DEFAULT", "cpu_threshold", fallback=65),
                    progress_callback=lambda cur, total: self.root.after(0, self.update_progress, cur, total),
                    result_callback=lambda result, cur, total: self.root.after(0, self.realtime_update_ui, result, cur, total)
                )
            else:
                # 处理单张图片
                recognized_text = FileProcessor.extract_text_from_image(
                    image_path=image_path,
                    ocr_engine=self.ocr_engine,
                    image_processor=self.image_processor,
                    optimize_mode=optimize_mode
                )
                result = {
                    "页码": 1,
                    "文本": recognized_text,
                    "状态": "未启用发票识别" if not self.invoice_recognition_var.get() else "待验证"
                }
                self.root.after(0, self.realtime_update_ui, result, 1, 1)
                results = [result]
            print("end ocr", time.time() - nowtimer)
            # 处理完成后显示总提示
            self.root.after(0, self.show_completion_message, len(results))

        except Exception as e:
            error_msg = f"{languages[self.current_language]['recognition_failed']}: {str(e)}"
            logging.error(error_msg, exc_info=True)
            self.root.after(0, self.update_ui_error, error_msg)

        finally:
            # 恢复 UI 状态
            self.root.after(0, self.stop_loading_animation)
            self.root.after(0, self.toggle_buttons_state, False)
            self.root.after(0, self.progress.stop)

    def show_completion_message(self, total):
        # 确保所有警告已显示
        self.root.after(500, lambda:
        messagebox.showinfo(
            languages[self.current_language]["success"],
            languages[self.current_language]["recognition_complete"].format(total=total)
        ))
    def update_ui_error(self, error_msg):
        # Stop progress bar
        self.progress.stop()
        # Enable buttons
        self.toggle_buttons_state(disabled=False)
        tk.messagebox.showerror(languages[self.current_language]["error"],
                                languages[self.current_language]["recognition_failed"] + f"\n{error_msg}")

    def toggle_buttons_state(self, disabled=True):
        """统一控制按钮禁用状态"""
        state = "disabled" if disabled else "normal"
        widgets = [
            self.start_button,
            self.browse_button,
            self.export_button,
            self.export_pos_button,
            self.ocr_module,
            self.optimize_mode
        ]
        for widget in widgets:
            widget.config(state=state)
    def write_to_excel(self):
        # Read information
        output_name = self.excel_name.get()
        data = self.file_characters.get("1.0", "end-1c")
        if not output_name:
            tk.messagebox.showerror(languages[self.current_language]["error"],
                                    languages[self.current_language]["empty_excel_name"])
            return

        try:
            self.file_processor.write_to_excel(data, output_name + '.xlsx')
            tk.messagebox.showinfo(languages[self.current_language]["success"],
                                   languages[self.current_language]["excel_saved"])
        except Exception as e:
            tk.messagebox.showerror(languages[self.current_language]["error"],
                                    languages[self.current_language]["excel_save_failed"].format(error=str(e)))


    def on_close(self):
        # Terminate main window
        self.root.destroy()


    def convert_pdf_to_image(self):
        pdf_path = self.pdf_to_image_path.get()
        output_format = self.output_image_format.get().lower()
        output_path = os.path.splitext(pdf_path)[0] + '_output'
        if not os.path.exists(output_path):
            os.makedirs(output_path)

            try:
                pdf_doc = fitz.open(pdf_path)
                for page_num in range(len(pdf_doc)):
                    page = pdf_doc.load_page(page_num)
                    pix = page.get_pixmap(dpi=300)  # Set DPI to 300
                    if output_format == "jpeg":
                        output_filename = os.path.join(output_path, f"page_{page_num + 1}.jpg")
                    elif output_format == "tiff":
                        output_filename = os.path.join(output_path, f"page_{page_num + 1}.tif")
                    else:
                        output_filename = os.path.join(output_path, f"page_{page_num + 1}.png")
                    pix.save(output_filename)

                message = languages[self.current_language]["pdf_converted"].format(
                    format=output_format.upper(), path=output_path)
                self.update_display_text(message)
                tk.messagebox.showinfo(languages[self.current_language]["success"],
                                       languages[self.current_language]["conversion_complete"].format(
                                           type="PDF→" + output_format.upper()))
            except Exception as e:
                message = languages[self.current_language]["conversion_failed"].format(
                    type="PDF→" + output_format.upper(), error=str(e))
                self.update_display_text(message, is_error=True)
                tk.messagebox.showerror(languages[self.current_language]["error"], message)

    def convert_image_to_pdf(self):
        image_paths = self.image_to_pdf_path.get().split(', ')
        if not image_paths or not image_paths[0]:
            tk.messagebox.showinfo(languages[self.current_language]["info"],
                                   languages[self.current_language]["no_image_selected"])
            return

        output_pdf_path = filedialog.asksaveasfilename(defaultextension=".pdf",
                                                       filetypes=[("PDF", "*.pdf"),
                                                                  (languages[self.current_language]["all_files"],
                                                                   "*.*")])
        if not output_pdf_path:
            return  # User canceled save

        try:
            images = []
            for image_path in image_paths:
                image = Image.open(image_path)
                pdf_page = image.convert('RGB')
                images.append(pdf_page)
            if images:
                images[0].save(output_pdf_path, save_all=True, append_images=images[1:])

            message = languages[self.current_language]["images_converted"].format(
                count=len(images), path=output_pdf_path)
            self.update_display_text(message)
            tk.messagebox.showinfo(languages[self.current_language]["success"],
                                   languages[self.current_language]["conversion_complete"].format(
                                       type=languages[self.current_language]["image_to_pdf"]))
        except Exception as e:
            message = languages[self.current_language]["conversion_failed"].format(
                type=languages[self.current_language]["image_to_pdf"], error=str(e))
            self.update_display_text(message, is_error=True)
            tk.messagebox.showerror(languages[self.current_language]["error"], message)

    def convert_txt_to_pdf(self):
        txt_path = self.txt_to_pdf_path.get()
        output_pdf_path = filedialog.asksaveasfilename(defaultextension=".pdf",
                                                       filetypes=[("PDF", "*.pdf"),
                                                                  (languages[self.current_language]["all_files"],
                                                                   "*.*")])
        if not output_pdf_path:
            return  # User canceled save

        try:
            with open(txt_path, 'r', encoding='utf-8') as file:
                text = file.read()
            c = canvas.Canvas(output_pdf_path, pagesize=letter)
            width, height = letter
            c.setFont("Helvetica", 12)
            text_lines = text.split('\n')
            y = height - 50
            for line in text_lines:
                c.drawString(50, y, line)
                y -= 20
                if y <= 50:
                    c.showPage()
                    c.setFont("Helvetica", 12)
                    y = height - 50
            c.save()

            message = languages[self.current_language]["txt_converted"].format(path=output_pdf_path)
            self.update_display_text(message)
            tk.messagebox.showinfo(languages[self.current_language]["success"],
                                   languages[self.current_language]["conversion_complete"].format(
                                       type=languages[self.current_language]["txt_to_pdf"]))
        except Exception as e:
            message = languages[self.current_language]["conversion_failed"].format(
                type=languages[self.current_language]["txt_to_pdf"], error=str(e))
            self.update_display_text(message, is_error=True)
            tk.messagebox.showerror(languages[self.current_language]["error"], message)

    def update_display_text(self, message, is_error=False):
        self.display_text.config(state='normal')
        self.display_text.insert(tk.END, message + '\n')
        self.display_text.config(state='disabled')
        if is_error:
            self.display_text.tag_add("error", "1.0", tk.END)
            self.display_text.tag_config("error", foreground="red")
        else:
            self.display_text.tag_add("success", "1.0", tk.END)
            self.display_text.tag_config("success", foreground="green")
        self.display_text.see(tk.END)

    def merge_pdfs(self):
        merged_pdf_path = filedialog.asksaveasfilename(defaultextension=".pdf",
                                                       filetypes=[("PDF", "*.pdf"),
                                                                  (languages[self.current_language]["all_files"],
                                                                   "*.*")])
        if not merged_pdf_path:
            return  # User canceled save

        try:
            from PyPDF2 import PdfMerger  # Updated to PdfMerger
            merger = PdfMerger()
            for pdf_file in self.pdf_files:
                merger.append(pdf_file)
            merger.write(merged_pdf_path)
            merger.close()

            message = languages[self.current_language]["pdf_merged"].format(path=merged_pdf_path)
            self.update_display_text(message)
            tk.messagebox.showinfo(languages[self.current_language]["success"],
                                   languages[self.current_language]["conversion_complete"].format(
                                       type=languages[self.current_language]["merge_pdf"]))
        except Exception as e:
            message = languages[self.current_language]["merge_failed"].format(error=str(e))
            self.update_display_text(message, is_error=True)
            tk.messagebox.showerror(languages[self.current_language]["error"],
                                    languages[self.current_language]["conversion_failed"].format(
                                        type=languages[self.current_language]["merge_pdf"], error=str(e)))

    def check_for_updates(self):
        update = UpdateManager(latest_version, self.current_language)
        update.check_update()
# 新增CPU监控类
class CPUMonitor:
    def __init__(self, threshold=None, window_size=5):
        self.threshold = threshold if threshold is not None else 65
        self.window_size = window_size
        self.cpu_history = []
        self.sleep_time = 0.01
        self.sensitivity = 0.3
        self.memory_threshold = 0.9  # 内存使用率阈值(90%)
        self.lock = threading.Lock()

    def adjust_speed(self):
        # 自动管理锁
        with self.lock:  # 这个with语句保证锁会被正确释放
            current_cpu = psutil.cpu_percent(interval=0.5)
            if len(self.cpu_history) >= self.window_size:
                self.cpu_history.pop(0)
            self.cpu_history.append(current_cpu)

            avg_cpu = sum(self.cpu_history) / len(self.cpu_history)
            if current_cpu > self.threshold:
                self.sleep_time = min(0.5, self.sleep_time * 1.2)  # 负载高时增加休眠
            else:
                self.sleep_time = max(0.01, self.sleep_time * 0.8)  # 负载低时减少休眠
            """调整处理速度，考虑CPU和内存"""
            if self.check_memory():
                self.sleep_time = min(1.0, self.sleep_time * 1.5)  # 内存不足时增加休眠
                return
    def check_memory(self):
        """检查内存使用情况"""
        mem = psutil.virtual_memory()
        if mem.percent > self.memory_threshold * 100:
            return True  # 内存不足
        return False
    def get_throttle_time(self):
        with self.lock:  # 添加获取读锁
            return self.sleep_time

    def get_current_cpu(self):
        return psutil.cpu_percent(interval=0.1)

class UpdateManager:
    def __init__(self, current_version, current_language):
        self.repo_url = "https://gitee.com/api/v5/repos/kris-1101/image-processing-tools/releases"
        self.token = "c586e921ff3515765a43129761dc0b5e"
        self.current_version = current_version
        self.current_language = current_language
        self.version_pattern = re.compile(r'^Image-tools_v(\d+\.\d+)(_part(\d+))?\.(zip|rar)$', re.IGNORECASE)

    def check_update(self):
        """Main update check method"""
        try:
            # Get all repository release information
            releases = self._get_releases()

            # Parse valid versions
            version_map = self._parse_releases(releases)

            if not version_map:
                messagebox.showinfo(languages[self.current_language]["no_update"],
                                    languages[self.current_language]["no_valid_version"])
                return

            # Get latest version
            latest_version = max(version_map.keys(), key=parse_version)

            if parse_version(latest_version) <= parse_version(self.current_version):
                messagebox.showinfo(languages[self.current_language]["no_update"],
                                    languages[self.current_language]["already_latest"].format(
                                        version=self.current_version))
                return

            # Handle volume download
            self._handle_download(version_map[latest_version], latest_version)

        except requests.RequestException as e:
            messagebox.showerror(languages[self.current_language]["network_error"],
                                 languages[self.current_language]["network_error"].format(error=str(e)))
        except Exception as e:
            messagebox.showerror(languages[self.current_language]["update_error"],
                                 languages[self.current_language]["update_error"].format(error=str(e)))

    def _get_releases(self):
        """Get repository release information"""
        headers = {
            'Accept': 'application/vnd.gitee+json',
            'Authorization': f'token {self.token}'
        }
        response = requests.get(self.repo_url, headers=headers, timeout=10)
        response.raise_for_status()
        return response.json()

    def _parse_releases(self, releases):
        """Parse release information and build version mapping"""
        version_map = defaultdict(list)

        # Define more flexible regex to match various possible volume file formats
        version_pattern = re.compile(
            r'^Image-tools_v(?P<version>\d+\.\d+)(\.zip\.00(?P<part>\d+))?\.zip$',
            re.IGNORECASE
        )

        for release in releases:
            # Only process stable releases
            if release.get("prerelease", True):
                continue

            for asset in release.get("assets", []):
                # Match version and volume information
                match = version_pattern.match(asset["name"] + '.zip')
                if not match:
                    continue

                # Extract version information
                base_version = match.group("version")
                part_num = int(match.group("part")) if match.group("part") else 0

                # Record volume file
                version_map[base_version].append({
                    "part": part_num,
                    "url": asset["browser_download_url"],
                    "name": asset["name"]
                })

        # Sort by volume number
        for ver in version_map:
            version_map[ver].sort(key=lambda x: x["part"])
        return version_map

    def _validate_version(self, version_str):
        """Validate version format validity"""
        try:
            parts = list(map(int, version_str.split(".")))
            return len(parts) == 2 and all(p >= 0 for p in parts)
        except ValueError:
            return False

    def _handle_download(self, assets, version):
        """Handle volume download process"""
        if not assets:
            messagebox.showwarning(languages[self.current_language]["update_error"],
                                   languages[self.current_language]["no_valid_parts"])
            return

        # Confirm download
        if not messagebox.askyesno(languages[self.current_language]["update_available"].format(version=version),
                                   languages[self.current_language]["update_available_tips"].format(
                                       version=version, count=len(assets))):
            return

        # Create download directory
        download_dir = Path(f"Image-tools_{version.replace('.', '_')}")
        download_dir.mkdir(exist_ok=True)

        # Download all volumes
        downloaded_files = []
        for idx, asset in enumerate(assets, 1):
            save_path = download_dir / asset["name"]
            if self._download_file(asset["url"], save_path, idx, len(assets)):
                downloaded_files.append(save_path)

        # Merge volumes
        if downloaded_files:
            final_file = download_dir / f"Image-tools_v{version}_full.zip"
            self._merge_files(downloaded_files, final_file)
            messagebox.showinfo(languages[self.current_language]["success"],
                                languages[self.current_language]["download_complete"].format(path=final_file))

    def _download_file(self, url, save_path, current, total):
        """Volume download with progress display"""
        try:
            progress_window = tk.Toplevel()
            progress_window.title(
                languages[self.current_language]["download_part"].format(current=current, total=total))

            # Progress components
            label = tk.Label(progress_window,
                             text=languages[self.current_language]["downloading"].format(name=save_path.name))
            label.pack(pady=5)

            progress = ttk.Progressbar(progress_window, length=300, mode='determinate')
            progress.pack(pady=5)

            response = requests.get(url, stream=True)
            total_size = int(response.headers.get('content-length', 1))
            downloaded = 0

            with open(save_path, 'wb') as f:
                for chunk in response.iter_content(chunk_size=8192):
                    if chunk:
                        f.write(chunk)
                        downloaded += len(chunk)
                        progress['value'] = (downloaded / total_size) * 100
                        progress_window.update()

            progress_window.destroy()
            return True
        except Exception as e:
            messagebox.showerror(languages[self.current_language]["download_failed"],
                                 languages[self.current_language]["download_failed"].format(name=save_path.name,
                                                                                            error=str(e)))
            return False

    def _merge_files(self, file_list, output_path):
        """Merge volume files"""
        try:
            with open(output_path, 'wb') as outfile:
                for file in sorted(file_list, key=lambda x: x.name):
                    with open(file, 'rb') as infile:
                        shutil.copyfileobj(infile, outfile)
            return True
        except Exception as e:
            messagebox.showerror(languages[self.current_language]["merge_failed"],
                                 languages[self.current_language]["merge_failed"].format(error=str(e)))
            return False

if __name__ == "__main__":
    print(time.time()-nowtimer)
    try:
        root = tk.Tk()
        app = ClassMain(root)
        root.mainloop()
    except Exception as e:
        logging.error(f"Application failed to start: {str(e)}")
        messagebox.showerror("Fatal Error", f"Application failed to start:\n{str(e)}")